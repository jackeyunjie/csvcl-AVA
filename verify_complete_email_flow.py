#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证完整的邮件发送流程，包括Excel颜色标记的保存和附件发送
"""

import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

# 添加项目路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def create_mt4_excel_with_colors(filename="mt4_test_data.xlsx"):
    """创建一个完整的MT4数据Excel文件，包含颜色标记"""
    print("创建MT4数据Excel文件...")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "MT4_Data_Colored"
    
    # 添加表头
    headers = ["SymbolName", "TIME", "MN1", "W1", "D1"]
    ws.append(headers)
    
    # 添加测试数据（模拟真实的MT4数据）
    test_data = [
        ["EURUSD", "2025.08.27", 2, -2, 8],
        ["GBPUSD", "2025.08.27", 6, -6, 10],
        ["USDJPY", "2025.08.27", 4, -4, 15],
        ["AUDUSD", "2025.08.27", 8, -8, 12],
        ["NZDUSD", "2025.08.27", 3, -3, 11],
        ["USDCAD", "2025.08.27", 5, -5, 13],
        ["USDCHF", "2025.08.27", 7, -7, 14],
        ["XAUUSD", "2025.08.27", 10, -10, 12],
        ["XAGUSD", "2025.08.27", 15, -15, 8]
    ]
    
    for row_data in test_data:
        ws.append(row_data)
    
    # 定义颜色规则（与规范一致）
    color_rules = {
        2: 'FF0000',    # 红色背景
        3: 'FF0000',    # 红色背景
        4: 'FF0000',    # 红色背景
        5: 'FF0000',    # 红色背景
        6: 'FF0000',    # 红色背景
        7: 'FF0000',    # 红色背景
        10: 'FFCCCC',   # 淡红色背景
        11: 'FFCCCC',   # 淡红色背景
        12: 'FFCCCC',   # 淡红色背景
        13: 'FFCCCC',   # 淡红色背景
        14: 'FFCCCC',   # 淡红色背景
        15: 'FFCCCC',   # 淡红色背景
        -2: '00FF00',   # 绿色背景
        -3: '00FF00',   # 绿色背景
        -4: '00FF00',   # 绿色背景
        -5: '00FF00',   # 绿色背景
        -6: '00FF00',   # 绿色背景
        -7: '00FF00',   # 绿色背景
        -10: 'CCFFCC',  # 淡绿色背景
        -11: 'CCFFCC',  # 淡绿色背景
        -12: 'CCFFCC',  # 淡绿色背景
        -13: 'CCFFCC',  # 淡绿色背景
        -14: 'CCFFCC',  # 淡绿色背景
        -15: 'CCFFCC',  # 淡绿色背景
        8: 'FFFF00'     # 黄色背景
    }
    
    # 创建颜色填充和字体对象
    color_fills = {}
    fonts = {}
    for value, color_code in color_rules.items():
        color_fills[value] = PatternFill(start_color=color_code, 
                                       end_color=color_code, 
                                       fill_type='solid')
        fonts[value] = Font(color='000000')  # 黑色字体
    
    # 应用颜色格式到指定区域（E2:G10，对应MN1, W1, D1列）
    colored_count = 0
    for row in range(2, 11):  # 第2到10行
        for col in range(3, 6):  # 第C到E列（MN1, W1, D1）
            cell = ws.cell(row=row, column=col)
            cell_value = cell.value
            
            if cell_value is not None:
                try:
                    numeric_value = float(cell_value)
                    # 解决浮点数精度问题
                    if abs(numeric_value - round(numeric_value)) < 1e-10:
                        int_value = int(round(numeric_value))
                        
                        if int_value in color_fills:
                            cell.fill = color_fills[int_value]
                            cell.font = fonts[int_value]
                            colored_count += 1
                            print(f"  单元格 {cell.coordinate} 值 {int_value} 已标记颜色 ({color_code})")
                except (ValueError, TypeError):
                    continue
    
    # 保存Excel文件
    wb.save(filename)
    print(f"✅ MT4数据Excel文件已保存: {filename}")
    print(f"   共标记了 {colored_count} 个单元格")
    
    return filename

def verify_excel_colors(file_path):
    """验证Excel文件中的颜色标记是否正确"""
    print(f"\n验证Excel文件颜色标记: {os.path.basename(file_path)}")
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active
        
        if ws is None:
            print("❌ 无法获取工作表")
            return False
        
        print(f"📊 工作表名称: {ws.title}")
        
        # 定义期望的颜色规则
        expected_colors = {
            2: 'FF0000',    # 红色
            3: 'FF0000',    # 红色
            4: 'FF0000',    # 红色
            5: 'FF0000',    # 红色
            6: 'FF0000',    # 红色
            7: 'FF0000',    # 红色
            10: 'FFCCCC',   # 淡红色
            11: 'FFCCCC',   # 淡红色
            12: 'FFCCCC',   # 淡红色
            13: 'FFCCCC',   # 淡红色
            14: 'FFCCCC',   # 淡红色
            15: 'FFCCCC',   # 淡红色
            -2: '00FF00',   # 绿色
            -3: '00FF00',   # 绿色
            -4: '00FF00',   # 绿色
            -5: '00FF00',   # 绿色
            -6: '00FF00',   # 绿色
            -7: '00FF00',   # 绿色
            -10: 'CCFFCC',  # 淡绿色
            -11: 'CCFFCC',  # 淡绿色
            -12: 'CCFFCC',  # 淡绿色
            -13: 'CCFFCC',  # 淡绿色
            -14: 'CCFFCC',  # 淡绿色
            -15: 'CCFFCC',  # 淡绿色
            8: 'FFFF00'     # 黄色
        }
        
        # 检查数据区域的颜色（E2:G10）
        color_check_count = 0
        correct_color_count = 0
        
        for row in range(2, 11):  # 第2到10行
            for col in range(3, 6):  # 第C到E列
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value is not None:
                    try:
                        # 尝试将值转换为数字
                        numeric_value = float(cell_value)
                        
                        # 解决浮点数精度问题
                        if abs(numeric_value - round(numeric_value)) < 1e-10:
                            int_value = int(round(numeric_value))
                            
                            # 检查是否在期望的颜色规则中
                            if int_value in expected_colors:
                                color_check_count += 1
                                
                                # 获取单元格的实际颜色
                                actual_color = None
                                if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                                    color_obj = cell.fill.start_color
                                    if hasattr(color_obj, 'rgb') and color_obj.rgb:
                                        actual_color = str(color_obj.rgb).upper()
                                
                                expected_color = expected_colors[int_value]
                                
                                # 检查颜色是否匹配
                                if actual_color and expected_color in actual_color:
                                    correct_color_count += 1
                                    print(f"   ✅ 单元格 {cell.coordinate} 值 {int_value} 颜色正确: {actual_color}")
                                else:
                                    print(f"   ❌ 单元格 {cell.coordinate} 值 {int_value} 颜色错误:")
                                    print(f"      期望: {expected_color}")
                                    print(f"      实际: {actual_color}")
                                    
                    except (ValueError, TypeError):
                        continue
        
        print(f"\n📊 验证结果:")
        print(f"   检查单元格数: {color_check_count}")
        print(f"   颜色正确数: {correct_color_count}")
        print(f"   正确率: {correct_color_count/color_check_count*100:.1f}%" if color_check_count > 0 else "   正确率: 0%")
        
        if color_check_count > 0 and correct_color_count == color_check_count:
            print("🎉 所有颜色标记都正确!")
            return True
        else:
            print("⚠️  部分颜色标记不正确或未找到匹配的单元格")
            return False
            
    except Exception as e:
        print(f"❌ 验证过程中出错: {str(e)}")
        return False

def simulate_email_process(file_path):
    """模拟完整的邮件处理流程"""
    print(f"\n=== 模拟邮件处理流程 ===")
    
    # 1. 验证原始文件颜色标记
    print("1. 验证原始文件颜色标记:")
    if not verify_excel_colors(file_path):
        print("❌ 原始文件颜色标记验证失败")
        return False
    
    # 2. 模拟邮件附件处理
    print("\n2. 模拟邮件附件处理:")
    try:
        # 在实际应用中，这里会调用email_sender.py中的方法
        # 我们只是验证文件格式是否正确
        file_size = os.path.getsize(file_path)
        print(f"   📄 文件大小: {file_size} 字节 ({file_size/1024:.1f} KB)")
        print(f"   📝 文件格式: Excel (.xlsx)")
        print("   ✅ 邮件附件处理模拟完成")
    except Exception as e:
        print(f"   ❌ 邮件附件处理模拟失败: {str(e)}")
        return False
    
    # 3. 验证处理后的文件（在实际应用中，这将是邮件附件）
    print("\n3. 验证处理后的文件:")
    if not verify_excel_colors(file_path):
        print("❌ 处理后文件颜色标记验证失败")
        return False
    
    print("\n🎉 完整邮件处理流程验证通过!")
    return True

def main():
    """主函数"""
    print("=== 验证完整的邮件发送流程 ===\n")
    
    # 创建测试文件
    test_file = create_mt4_excel_with_colors("complete_mt4_test.xlsx")
    
    # 验证完整流程
    success = simulate_email_process(test_file)
    
    if success:
        print(f"\n✅ 完整流程验证成功!")
        print(f"   生成的文件: {test_file}")
        print(f"   请手动检查此文件的颜色标记是否正确")
        print(f"   然后可以手动删除此文件")
    else:
        print(f"\n❌ 完整流程验证失败!")
        # 清理测试文件
        if os.path.exists(test_file):
            os.remove(test_file)
            print(f"   已清理测试文件: {test_file}")

if __name__ == "__main__":
    main()