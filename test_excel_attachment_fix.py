#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证邮件附件修复 - 测试脚本
用于确认Excel附件能否正确保留颜色标记
"""

import os
import sys
import time
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

# 添加项目路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from email_sender import EmailSender

def create_test_excel_with_colors(filename="test_email_excel.xlsx"):
    """创建一个带颜色标记的测试Excel文件"""
    print("创建带颜色标记的测试Excel文件...")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "MT4_Data_Test"
    
    # 添加表头
    headers = ["SymbolName", "TIME", "MN1", "W1", "D1"]
    ws.append(headers)
    
    # 添加测试数据
    test_data = [
        ["EURUSD", "2025.08.27", 2, -2, 8],
        ["GBPUSD", "2025.08.27", 6, -6, 10],
        ["USDJPY", "2025.08.27", 4, -4, 15],
        ["AUDUSD", "2025.08.27", 8, -8, 12],
        ["NZDUSD", "2025.08.27", 3, -3, 11]
    ]
    
    for row_data in test_data:
        ws.append(row_data)
    
    # 定义颜色规则
    color_rules = {
        2: 'FF0000',    # 红色背景
        3: 'FF0000',    # 红色背景
        4: 'FF0000',    # 红色背景
        5: 'FF0000',    # 红色背景
        6: 'FF0000',    # 红色背景
        7: 'FF0000',    # 红色背景
        10: 'FFCCCC',   # 淡红色背景
        11: 'FFCCCC',   # 淡红色背景
        12: 'FFCCCC',   # 淡红色背景
        13: 'FFCCCC',   # 淡红色背景
        14: 'FFCCCC',   # 淡红色背景
        15: 'FFCCCC',   # 淡红色背景
        -2: '00FF00',   # 绿色背景
        -3: '00FF00',   # 绿色背景
        -4: '00FF00',   # 绿色背景
        -5: '00FF00',   # 绿色背景
        -6: '00FF00',   # 绿色背景
        -7: '00FF00',   # 绿色背景
        -10: 'CCFFCC',  # 淡绿色背景
        -11: 'CCFFCC',  # 淡绿色背景
        -12: 'CCFFCC',  # 淡绿色背景
        -13: 'CCFFCC',  # 淡绿色背景
        -14: 'CCFFCC',  # 淡绿色背景
        -15: 'CCFFCC',  # 淡绿色背景
        8: 'FFFF00'     # 黄色背景
    }
    
    # 创建颜色填充和字体对象
    color_fills = {}
    fonts = {}
    for value, color_code in color_rules.items():
        color_fills[value] = PatternFill(start_color=color_code, 
                                       end_color=color_code, 
                                       fill_type='solid')
        fonts[value] = Font(color='000000')  # 黑色字体
    
    # 应用颜色格式到指定区域（E2:G6，对应MN1, W1, D1列）
    colored_count = 0
    for row in range(2, 7):  # 第2到6行
        for col in range(3, 6):  # 第C到E列（MN1, W1, D1）
            cell = ws.cell(row=row, column=col)
            cell_value = cell.value
            
            if cell_value is not None:
                try:
                    numeric_value = float(cell_value)
                    # 解决浮点数精度问题
                    if abs(numeric_value - round(numeric_value)) < 1e-10:
                        int_value = int(round(numeric_value))
                        
                        if int_value in color_fills:
                            cell.fill = color_fills[int_value]
                            cell.font = fonts[int_value]
                            colored_count += 1
                            print(f"  单元格 {cell.coordinate} 值 {int_value} 已标记颜色 ({color_rules[int_value]})")
                except (ValueError, TypeError):
                    continue
    
    # 调整B列宽度为其他列的1.5倍
    ws.column_dimensions['B'].width = 20  # 默认宽度约为13
    
    # 保存Excel文件
    wb.save(filename)
    print(f"✅ 测试Excel文件已保存: {filename}")
    print(f"   共标记了 {colored_count} 个单元格")
    
    return filename

def send_test_email(excel_file):
    """发送测试邮件"""
    print("\n=== 发送测试邮件 ===")
    
    # 创建邮件发送器
    sender = EmailSender()
    
    # 检查配置
    if not sender.username or not sender.password:
        print("❌ 邮箱配置不完整，请检查email_config.ini文件")
        return False
    
    # 检查文件是否存在
    if not os.path.exists(excel_file):
        print(f"❌ 测试文件不存在: {excel_file}")
        return False
    
    # 定义测试收件人（请修改为实际的邮箱地址）
    recipients = ["447372703@qq.com"]  # 可以修改为其他测试邮箱
    
    # 创建邮件主题
    subject = f"MT4数据处理测试报告 - 颜色标记验证 {time.strftime('%Y-%m-%d %H:%M:%S')}"
    
    # 发送邮件
    print(f"📧 准备发送测试邮件...")
    print(f"   发件人: {sender.username}")
    print(f"   收件人: {', '.join(recipients)}")
    print(f"   附件: {excel_file}")
    
    success = sender.send_mt4_report(
        excel_files=[excel_file],
        image_files=[],  # 不添加图片附件以简化测试
        recipients=recipients,
        subject=subject,
        data_changes=[]  # 不添加数据变化以简化测试
    )
    
    if success:
        print("✅ 测试邮件发送成功!")
        print("请检查收件箱，验证邮件附件中的Excel文件颜色标记是否正确显示")
        return True
    else:
        print("❌ 测试邮件发送失败!")
        return False

def main():
    """主函数"""
    print("=== 邮件附件Excel颜色标记测试 ===")
    
    # 1. 创建测试Excel文件
    excel_file = create_test_excel_with_colors()
    
    # 2. 发送测试邮件
    success = send_test_email(excel_file)
    
    # 3. 提示用户检查结果
    if success:
        print("\n✅ 测试完成")
        print("请检查邮箱收件箱，确认以下几点:")
        print("1. 邮件是否成功接收")
        print("2. Excel附件是否可以正常打开")
        print("3. Excel文件中的颜色标记是否正确显示")
        print(f"4. 测试文件位置: {os.path.abspath(excel_file)}")
    else:
        print("\n❌ 测试失败")
        print("请检查上述错误信息并修复问题")

if __name__ == "__main__":
    main()