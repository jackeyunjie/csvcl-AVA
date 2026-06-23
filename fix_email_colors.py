#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修复邮件中颜色显示问题
"""

import os
import sys
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont, ImageEnhance
import pandas as pd
from openpyxl import load_workbook
from csv_color_marker import CSVColorMarker
from email_sender import EmailSender

def enhance_image_quality(image_path, output_path=None):
    """增强图片质量，确保颜色清晰可见"""
    print(f"增强图片质量: {image_path}")
    
    if not os.path.exists(image_path):
        print(f"❌ 文件不存在: {image_path}")
        return None
    
    try:
        # 打开原始图片
        img = Image.open(image_path)
        
        # 如果未指定输出路径，创建一个新路径
        if output_path is None:
            base_name = os.path.splitext(image_path)[0]
            output_path = f"{base_name}_enhanced.png"
        
        # 增强图片质量
        
        # 增强对比度
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(1.2)  # 增强20%
        
        # 增强饱和度
        enhancer = ImageEnhance.Color(img)
        img = enhancer.enhance(1.3)  # 增强30%
        
        # 增强亮度
        enhancer = ImageEnhance.Brightness(img)
        img = enhancer.enhance(1.1)  # 增强10%
        
        # 保存增强后的图片（高质量）
        img.save(output_path, format='PNG', quality=95, optimize=True, dpi=(300, 300))
        print(f"✅ 增强后的图片已保存: {output_path}")
        
        return output_path
    
    except Exception as e:
        print(f"❌ 增强图片时出错: {str(e)}")
        return None

def create_html_preview(excel_file, image_file):
    """创建HTML预览，模拟邮件中的显示效果"""
    print(f"创建HTML预览...")
    
    if not os.path.exists(excel_file) or not os.path.exists(image_file):
        print("❌ 缺少Excel文件或图片文件")
        return None
    
    try:
        # 创建HTML文件
        html_path = "email_preview.html"
        
        # 获取文件名
        excel_name = os.path.basename(excel_file)
        image_name = os.path.basename(image_file)
        
        # 创建HTML内容
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>邮件预览</title>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .header {{ background-color: #4CAF50; color: white; padding: 15px; text-align: center; }}
                .content {{ padding: 20px; }}
                .file-list {{ background-color: #f9f9f9; padding: 15px; margin: 10px 0; }}
                .image-container {{ margin: 20px 0; text-align: center; }}
                .image-container img {{ max-width: 100%; height: auto; border: 1px solid #ddd; }}
                .footer {{ background-color: #f5f5f5; padding: 10px; text-align: center; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>📊 MT4数据文件</h2>
                <p>{pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div class="content">
                <div class="file-list">
                    <h3>📊 Excel文件:</h3>
                    <ul>
                        <li>📊 {excel_name}</li>
                    </ul>
                </div>
                
                <h3>🖼️ 数据截图:</h3>
                <div class="image-container">
                    <h4>{image_name}</h4>
                    <img src="{image_file}" alt="{image_name}">
                </div>
            </div>
            
            <div class="footer">
                <p>🤖 此邮件由MT4数据处理系统自动发送</p>
            </div>
        </body>
        </html>
        """
        
        # 保存HTML文件
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"✅ HTML预览已创建: {html_path}")
        
        # 尝试打开HTML文件
        try:
            import webbrowser
            webbrowser.open(html_path)
            print("✅ HTML预览已在浏览器中打开")
        except Exception as e:
            print(f"❌ 无法自动打开HTML预览: {str(e)}")
        
        return html_path
    
    except Exception as e:
        print(f"❌ 创建HTML预览时出错: {str(e)}")
        return None

def patch_save_range_as_image():
    """修补CSVColorMarker类的save_range_as_image方法，提高颜色清晰度"""
    print("修补图片生成代码...")
    
    try:
        # 创建一个增强版的save_range_as_image方法
        def enhanced_save_range_as_image(self, worksheet, range_str="A1:M40", output_path=None, include_headers=True):
            """
            增强版的save_range_as_image方法，提高颜色清晰度
            """
            try:
                print(f"正在将区域 {range_str} 保存为高质量图片（包含行列号）...")
                
                # 解析区域
                start_col, start_row, end_col, end_row = self.parse_range(range_str)
                
                # 设置单元格和图片的基本尺寸（增加尺寸以提高清晰度）
                cell_width = 100  # 原来是80
                cell_height = 30  # 原来是25
                header_width = 50  # 原来是40
                header_height = 30  # 原来是25
                
                # 计算图片尺寸
                if include_headers:
                    img_width = header_width + (end_col - start_col + 1) * cell_width
                    img_height = header_height + (end_row - start_row + 1) * cell_height
                else:
                    img_width = (end_col - start_col + 1) * cell_width
                    img_height = (end_row - start_row + 1) * cell_height
                
                # 创建图片（使用RGBA模式以支持透明度）
                img = Image.new('RGBA', (img_width, img_height), (255, 255, 255, 255))
                draw = ImageDraw.Draw(img)
                
                # 尝试加载更好的字体
                try:
                    font = ImageFont.truetype("arial.ttf", 12)  # 增大字体
                    header_font = ImageFont.truetype("arial.ttf", 11)
                except:
                    try:
                        font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 12)
                        header_font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 11)
                    except:
                        font = ImageFont.load_default()
                        header_font = ImageFont.load_default()
                
                # 定义增强的颜色映射（更鲜明的颜色）
                color_map = {
                    # 红色系列（更鲜明）
                    '00FF0000': (255, 0, 0, 255),    # 红色背景
                    'FF0000': (255, 0, 0, 255),      # 红色背景
                    
                    # 淡红色系列（更鲜明）
                    '00FFCCCC': (255, 180, 180, 255), # 淡红色背景
                    'FFCCCC': (255, 180, 180, 255),   # 淡红色背景
                    
                    # 绿色系列（更鲜明）
                    '0000FF00': (0, 255, 0, 255),    # 绿色背景
                    '00FF00': (0, 255, 0, 255),      # 绿色背景
                    
                    # 淡绿色系列（更鲜明）
                    '00CCFFCC': (180, 255, 180, 255), # 淡绿色背景
                    'CCFFCC': (180, 255, 180, 255),   # 淡绿色背景
                    
                    # 黄色系列（更鲜明）
                    '00FFFF00': (255, 255, 0, 255),  # 黄色背景
                    'FFFF00': (255, 255, 0, 255),    # 黄色背景
                    
                    # 默认颜色
                    'FFFFFF': (255, 255, 255, 255),  # 白色背景
                    '00FFFFFF': (255, 255, 255, 255), # 白色背景
                    '00000000': (255, 255, 255, 255)  # 透明色作为白色处理
                }
                
                # 绘制行号和列号的偏移量
                offset_x = header_width if include_headers else 0
                offset_y = header_height if include_headers else 0
                
                # 绘制列号标题（如果包含标题）
                if include_headers:
                    # 绘制左上角空白区域
                    draw.rectangle([0, 0, header_width, header_height], 
                                 fill=(220, 220, 220, 255), outline=(0, 0, 0, 255))
                    
                    # 绘制列号
                    for col in range(start_col, end_col + 1):
                        x = offset_x + (col - start_col) * cell_width
                        y = 0
                        
                        # 绘制列号背景
                        draw.rectangle([x, y, x + cell_width, y + header_height], 
                                     fill=(220, 220, 220, 255), outline=(0, 0, 0, 255))
                        
                        # 绘制列号文本
                        col_letter = self.get_column_letter(col)
                        bbox = draw.textbbox((0, 0), col_letter, font=header_font)
                        text_width = bbox[2] - bbox[0]
                        text_height = bbox[3] - bbox[1]
                        text_x = x + (cell_width - text_width) // 2
                        text_y = y + (header_height - text_height) // 2
                        draw.text((text_x, text_y), col_letter, fill=(0, 0, 0, 255), font=header_font)
                    
                    # 绘制行号
                    for row in range(start_row, end_row + 1):
                        x = 0
                        y = offset_y + (row - start_row) * cell_height
                        
                        # 绘制行号背景
                        draw.rectangle([x, y, x + header_width, y + cell_height], 
                                     fill=(220, 220, 220, 255), outline=(0, 0, 0, 255))
                        
                        # 绘制行号文本
                        row_text = str(row)
                        bbox = draw.textbbox((0, 0), row_text, font=header_font)
                        text_width = bbox[2] - bbox[0]
                        text_height = bbox[3] - bbox[1]
                        text_x = x + (header_width - text_width) // 2
                        text_y = y + (cell_height - text_height) // 2
                        draw.text((text_x, text_y), row_text, fill=(0, 0, 0, 255), font=header_font)
                
                # 绘制数据单元格
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        # 计算单元格在图片中的位置
                        x = offset_x + (col - start_col) * cell_width
                        y = offset_y + (row - start_row) * cell_height
                        
                        # 获取单元格
                        cell = worksheet.cell(row=row, column=col)
                        cell_value = cell.value if cell.value is not None else ""
                        
                        # 确定背景颜色
                        bg_color = (255, 255, 255, 255)  # 默认白色
                        
                        if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                            # 获取颜色代码
                            color_obj = cell.fill.start_color
                            color_code = None
                            
                            # 尝试获取RGB值
                            if hasattr(color_obj, 'rgb') and color_obj.rgb:
                                color_code = str(color_obj.rgb)
                            
                            # 如果有颜色代码且不是透明或默认值
                            if color_code and color_code in color_map:
                                bg_color = color_map[color_code]
                                # 打印调试信息
                                print(f"单元格 {self.get_column_letter(col)}{row} 值={cell_value} 颜色代码={color_code} RGB={bg_color}")
                        
                        # 绘制单元格背景
                        draw.rectangle([x, y, x + cell_width, y + cell_height], 
                                     fill=bg_color, outline=(0, 0, 0, 255))
                        
                        # 绘制单元格文本
                        if cell_value is not None and str(cell_value).strip() != "":
                            text = str(cell_value)
                            # 计算文本位置（居中）
                            bbox = draw.textbbox((0, 0), text, font=font)
                            text_width = bbox[2] - bbox[0]
                            text_height = bbox[3] - bbox[1]
                            text_x = x + (cell_width - text_width) // 2
                            text_y = y + (cell_height - text_height) // 2
                            
                            # 绘制文本（黑色字体）
                            draw.text((text_x, text_y), text, fill=(0, 0, 0, 255), font=font)
                
                # 生成输出路径
                if output_path is None:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_path = os.path.join(os.path.dirname(__file__), f"excel_full_table_{timestamp}.png")
                
                # 增强图片质量
                enhancer = ImageEnhance.Contrast(img)
                img = enhancer.enhance(1.2)  # 增强对比度
                
                enhancer = ImageEnhance.Color(img)
                img = enhancer.enhance(1.3)  # 增强饱和度
                
                # 保存高质量图片
                img.save(output_path, 'PNG', quality=95, optimize=True, dpi=(300, 300))
                print(f"[OK] 已保存高质量表格图片: {os.path.basename(output_path)}")
                print(f"  完整路径: {os.path.abspath(output_path)}")
                print(f"  图片尺寸: {img_width} x {img_height} 像素")
                print(f"  包含行列号: {'是' if include_headers else '否'}")
                
                return output_path
                
            except Exception as e:
                print(f"[ERROR] 保存图片时出错: {str(e)}")
                return None
        
        # 替换CSVColorMarker类的save_range_as_image方法
        CSVColorMarker.save_range_as_image = enhanced_save_range_as_image
        print("✅ 已成功修补图片生成代码")
        return True
        
    except Exception as e:
        print(f"❌ 修补图片生成代码失败: {str(e)}")
        return False

def patch_email_sender():
    """修补EmailSender类的embed_image方法，确保图片以最佳质量嵌入"""
    print("修补邮件发送代码...")
    
    try:
        # 创建一个增强版的embed_image方法
        def enhanced_embed_image(self, msg, image_path, cid):
            """
            增强版的embed_image方法，确保图片以最佳质量嵌入
            """
            try:
                filename = os.path.basename(image_path)
                
                # 先增强图片质量
                enhanced_image_path = image_path
                if image_path.endswith('.png') and not image_path.endswith('_enhanced.png'):
                    try:
                        from PIL import Image, ImageEnhance
                        img = Image.open(image_path)
                        
                        # 增强对比度
                        enhancer = ImageEnhance.Contrast(img)
                        img = enhancer.enhance(1.2)
                        
                        # 增强饱和度
                        enhancer = ImageEnhance.Color(img)
                        img = enhancer.enhance(1.3)
                        
                        # 保存增强后的图片
                        enhanced_image_path = os.path.splitext(image_path)[0] + "_enhanced.png"
                        img.save(enhanced_image_path, format='PNG', quality=95, optimize=True, dpi=(300, 300))
                        print(f"✅ 已增强图片质量: {os.path.basename(enhanced_image_path)}")
                    except Exception as e:
                        print(f"⚠️ 增强图片质量失败，使用原图: {str(e)}")
                
                with open(enhanced_image_path, 'rb') as f:
                    img_data = f.read()
                    
                    # 根据文件扩展名确定图片类型
                    if filename.lower().endswith('.jpg') or filename.lower().endswith('.jpeg'):
                        image_type = 'jpeg'
                    elif filename.lower().endswith('.png'):
                        image_type = 'png'
                    else:
                        image_type = 'jpeg'
                    
                    # 使用高质量设置
                    image = MIMEImage(img_data, _subtype=image_type)
                    image.add_header('Content-ID', f'<{cid}>')
                    image.add_header('Content-Disposition', 'inline', filename=filename)
                    
                    # 添加额外的头信息以确保高质量显示
                    image.add_header('X-Apple-Content-Length', str(len(img_data)))
                    image.add_header('Content-Type', f'image/{image_type}; name="{filename}"')
                    
                    msg.attach(image)
                    print(f"✅ 已以高质量嵌入图片: {filename}")
                    
            except Exception as e:
                print(f"❌ 嵌入图片失败 {image_path}: {str(e)}")
        
        # 替换EmailSender类的embed_image方法
        from email.mime.image import MIMEImage
        EmailSender.embed_image = enhanced_embed_image
        print("✅ 已成功修补邮件发送代码")
        return True
        
    except Exception as e:
        print(f"❌ 修补邮件发送代码失败: {str(e)}")
        return False

def test_with_real_data():
    """使用真实数据测试颜色显示"""
    print("\n=== 使用真实数据测试 ===")
    
    # 查找最新的CSV文件
    import glob
    csv_files = glob.glob("*.csv")
    
    if not csv_files:
        print("❌ 未找到CSV文件")
        return False
    
    # 按修改时间排序，获取最新的文件
    latest_csv = max(csv_files, key=os.path.getmtime)
    print(f"使用最新CSV文件: {latest_csv}")
    
    try:
        # 应用修补
        patch_save_range_as_image()
        patch_email_sender()
        
        # 处理CSV文件
        marker = CSVColorMarker()
        excel_file = marker.process_csv_file(latest_csv)
        
        if not excel_file:
            print("❌ 处理CSV文件失败")
            return False
        
        # 查找生成的图片
        image_file = os.path.splitext(excel_file)[0].replace('_colored', '_full_table') + '.png'
        
        if not os.path.exists(image_file):
            print(f"❌ 未找到图片文件: {image_file}")
            return False
        
        # 创建HTML预览
        html_file = create_html_preview(excel_file, image_file)
        
        # 询问是否发送测试邮件
        send_email = input("\n是否发送测试邮件？(y/n): ").strip().lower()
        if send_email == 'y':
            # 获取收件人邮箱
            recipient = input("请输入收件人邮箱: ").strip()
            if not recipient:
                print("❌ 未提供收件人邮箱")
                return False
            
            # 发送测试邮件
            sender = EmailSender()
            result = sender.send_mt4_report(
                excel_files=[excel_file],
                image_files=[image_file],
                recipients=[recipient],
                subject="颜色标注修复测试邮件"
            )
            
            if result:
                print("✅ 测试邮件发送成功！请检查收件箱")
            else:
                print("❌ 测试邮件发送失败")
        
        return True
        
    except Exception as e:
        print(f"❌ 测试失败: {str(e)}")
        return False

def main():
    """主函数"""
    print("=== 邮件颜色显示修复工具 ===")
    print("此工具用于修复邮件中颜色标注不显示的问题")
    print("=" * 40)
    
    print("\n可用选项:")
    print("1. 修补图片生成代码")
    print("2. 修补邮件发送代码")
    print("3. 增强现有图片质量")
    print("4. 创建HTML预览")
    print("5. 使用真实数据测试")
    print("6. 全部修复并测试")
    print("0. 退出")
    
    choice = input("\n请选择操作 (0-6): ").strip()
    
    if choice == '1':
        patch_save_range_as_image()
    elif choice == '2':
        patch_email_sender()
    elif choice == '3':
        # 查找最新的图片
        import glob
        image_files = glob.glob("*_full_table.png")
        if image_files:
            latest_image = max(image_files, key=os.path.getmtime)
            enhance_image_quality(latest_image)
        else:
            print("❌ 未找到图片文件")
    elif choice == '4':
        # 查找最新的Excel和图片
        import glob
        excel_files = glob.glob("*_colored.xlsx")
        image_files = glob.glob("*_full_table.png")
        
        if excel_files and image_files:
            latest_excel = max(excel_files, key=os.path.getmtime)
            latest_image = max(image_files, key=os.path.getmtime)
            create_html_preview(latest_excel, latest_image)
        else:
            print("❌ 未找到Excel或图片文件")
    elif choice == '5':
        test_with_real_data()
    elif choice == '6':
        patch_save_range_as_image()
        patch_email_sender()
        test_with_real_data()
    elif choice == '0':
        print("退出程序")
    else:
        print("❌ 无效选择")

if __name__ == "__main__":
    main()