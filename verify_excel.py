#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件验证工具
用于检查生成的Excel文件是否可以正常读取
"""

import os
import sys
from openpyxl import load_workbook

def test_excel_file(file_path):
    """
    测试Excel文件是否可以正常打开和读取
    
    Args:
        file_path (str): Excel文件路径
    """
    try:
        print(f"正在测试文件: {file_path}")
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            print(f"❌ 文件不存在: {file_path}")
            return False
        
        # 检查文件大小
        file_size = os.path.getsize(file_path)
        print(f"📁 文件大小: {file_size} 字节")
        
        if file_size == 0:
            print("❌ 文件为空")
            return False
        
        # 尝试加载工作簿
        print("🔄 正在尝试打开Excel文件...")
        wb = load_workbook(file_path)
        
        print("✅ Excel文件加载成功!")
        
        # 获取工作表信息
        ws = wb.active
        print(f"📊 工作表名称: {ws.title}")
        print(f"📏 最大行数: {ws.max_row}")
        print(f"📏 最大列数: {ws.max_column}")
        
        # 检查一些单元格的内容
        print("\n🔍 检查单元格内容:")
        for row in range(1, min(6, ws.max_row + 1)):
            for col in range(1, min(6, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    print(f"  {chr(64+col)}{row}: {cell.value}")
        
        # 检查颜色格式
        print("\n🎨 检查颜色格式:")
        colored_cells = 0
        for row in range(2, min(11, ws.max_row + 1)):  # 检查前几行
            for col in range(5, 8):  # E, F, G列
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.index != '00000000':  # 不是默认颜色
                    colored_cells += 1
                    col_letter = chr(64+col)
                    print(f"  {col_letter}{row}: 值={cell.value}, 背景色={cell.fill.start_color.index}")
        
        print(f"🎯 发现 {colored_cells} 个有颜色的单元格")
        
        # 检查列宽
        print("\n📐 检查列宽:")
        for col in range(1, min(8, ws.max_column + 1)):
            col_letter = chr(64+col)
            width = ws.column_dimensions[col_letter].width
            print(f"  {col_letter}列宽度: {width}")
        
        wb.close()
        print("\n✅ 文件验证完成 - 文件格式正常!")
        return True
        
    except Exception as e:
        print(f"❌ 文件验证失败: {str(e)}")
        return False

def main():
    """主函数"""
    # 测试目录中的所有colored.xlsx文件
    current_dir = os.path.dirname(__file__)
    
    xlsx_files = [f for f in os.listdir(current_dir) if f.endswith('_colored.xlsx')]
    
    if not xlsx_files:
        print("❌ 未找到任何_colored.xlsx文件")
        return
    
    print(f"📂 找到 {len(xlsx_files)} 个Excel文件")
    print("=" * 50)
    
    all_valid = True
    for file_name in xlsx_files:
        file_path = os.path.join(current_dir, file_name)
        is_valid = test_excel_file(file_path)
        all_valid = all_valid and is_valid
        print("-" * 50)
    
    if all_valid:
        print("\n🎉 所有文件都验证通过!")
    else:
        print("\n⚠️ 部分文件存在问题")

if __name__ == "__main__":
    main()