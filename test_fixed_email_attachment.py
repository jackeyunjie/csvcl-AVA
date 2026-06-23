#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试修复后的邮件附件功能
"""

import os
import sys
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# 添加项目路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from email_sender import EmailSender

def create_test_excel_with_colors():
    """创建一个带颜色标记的测试Excel文件"""
    print("创建带颜色标记的测试Excel文件...")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "MT4_Data_Colored"
    
    # 添加表头
    headers = ["SymbolName", "TIME", "MN1", "W1", "D1"]
    ws.append(headers)
    
    # 添加一些测试数据
    test_data = [
        ["EURUSD", "2025.08.27", 2, -2, 8],
        ["GBPUSD", "2025.08.27", 6, -6, 10],
        ["USDJPY", "2025.08.27", 4, -4, 15],
        ["AUDUSD", "2025.08.27", 8, -8, 12]
    ]
    
    for row_data in test_data:
        ws.append(row_data)
    
    # 定义颜色规则（与process_real_mt4_data.py中的一致）
    color_rules = {
        2: 'FF0000',    # 红色
        3: 'FF0000',    # 红色
        4: 'FF0000',    # 红色
        5: 'FF0000',    # 红色
        6: 'FF0000',    # 红色
        7: 'FF0000',    # 红色
        10: 'FFCCCC',   # 淡红色
        11: 'FFCCCC',   # 淡红色
        12: 'FFCCCC',   # 淡红色
        13: 'FFCCCC',   # 淡红色
        14: 'FFCCCC',   # 淡红色
        15: 'FFCCCC',   # 淡红色
        -2: '00FF00',   # 绿色
        -3: '00FF00',   # 绿色
        -4: '00FF00',   # 绿色
        -5: '00FF00',   # 绿色
        -6: '00FF00',   # 绿色
        -7: '00FF00',   # 绿色
        -10: 'CCFFCC',  # 淡绿色
        -11: 'CCFFCC',  # 淡绿色
        -12: 'CCFFCC',  # 淡绿色
        -13: 'CCFFCC',  # 淡绿色
        -14: 'CCFFCC',  # 淡绿色
        -15: 'CCFFCC',  # 淡绿色
        8: 'FFFF00'     # 黄色
    }
    
    # 创建颜色填充对象
    color_fills = {}
    fonts = {}
    for value, color_code in color_rules.items():
        color_fills[value] = PatternFill(start_color=color_code, 
                                       end_color=color_code, 
                                       fill_type='solid')
        fonts[value] = Font(color='000000')  # 黑色字体
    
    # 应用颜色标记到数据区域（E2:G5）
    colored_count = 0
    for row in range(2, 6):  # 第2到5行
        for col in range(3, 6):  # 第C到E列（对应MN1, W1, D1）
            cell = ws.cell(row=row, column=col)
            cell_value = cell.value
            
            if cell_value is not None:
                try:
                    numeric_value = float(cell_value)
                    # 解决浮点数精度问题
                    if abs(numeric_value - round(numeric_value)) < 1e-10:
                        int_value = int(round(numeric_value))
                        
                        if int_value in color_fills:
                            cell.fill = color_fills[int_value]
                            cell.font = fonts[int_value]
                            colored_count += 1
                            print(f"  单元格 {cell.coordinate} 值 {int_value} 已标记颜色")
                except (ValueError, TypeError):
                    continue
    
    # 保存文件
    test_file = "test_mt4_colored.xlsx"
    wb.save(test_file)
    print(f"✅ 测试Excel文件已保存: {test_file}")
    print(f"   共标记了 {colored_count} 个单元格")
    
    return test_file

def test_email_attachment():
    """测试邮件附件功能"""
    print("\n=== 测试邮件附件功能 ===")
    
    # 创建测试Excel文件
    test_file = create_test_excel_with_colors()
    
    try:
        # 创建邮件发送器
        sender = EmailSender()
        
        # 创建测试邮件消息
        msg = MIMEMultipart()
        msg['From'] = 'test@example.com'
        msg['To'] = 'recipient@example.com'
        msg['Subject'] = '测试Excel附件颜色标记'
        
        # 使用修复后的attach_file方法添加附件
        sender.attach_file(msg, test_file)
        
        print("✅ 邮件附件添加成功")
        print("   请检查生成的邮件是否正确包含颜色标记")
        
        return test_file
        
    except Exception as e:
        print(f"❌ 邮件附件测试失败: {str(e)}")
        return None

def main():
    """主函数"""
    print("=== 测试修复后的邮件附件功能 ===\n")
    
    # 测试邮件附件功能
    test_file = test_email_attachment()
    
    # 清理测试文件
    if test_file and os.path.exists(test_file):
        # 暂时保留文件供手动检查
        print(f"\n📁 测试文件已生成: {test_file}")
        print("   请手动检查此文件的颜色标记是否正确")
        print("   然后可以手动删除此文件")
    
    print("\n🎉 邮件附件功能测试完成!")

if __name__ == "__main__":
    main()