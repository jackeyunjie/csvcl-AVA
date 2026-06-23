#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试Excel颜色标记在邮件发送过程中的保存情况
"""

import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header

def create_test_excel_with_colors():
    """创建一个带颜色标记的测试Excel文件"""
    print("创建带颜色标记的测试Excel文件...")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Test_Colors"
    
    # 添加一些测试数据
    data = [
        ["Symbol", "MN1", "W1", "D1"],
        ["EURUSD", 2, -2, 8],
        ["GBPUSD", 6, -6, 10],
        ["USDJPY", 4, -4, 15],
        ["AUDUSD", 8, -8, 12]
    ]
    
    # 写入数据
    for row in data:
        ws.append(row)
    
    # 定义颜色规则
    color_rules = {
        2: 'FF0000',    # 红色
        6: 'FF0000',    # 红色
        -2: '00FF00',   # 绿色
        -6: '00FF00',   # 绿色
        8: 'FFFF00',    # 黄色
        10: 'FFCCCC',   # 淡红色
        15: 'FFCCCC',   # 淡红色
        -8: 'CCFFCC',   # 淡绿色
        12: 'CCFFCC'    # 淡绿色
    }
    
    # 创建颜色填充对象
    color_fills = {}
    for value, color_code in color_rules.items():
        color_fills[value] = PatternFill(start_color=color_code, 
                                       end_color=color_code, 
                                       fill_type='solid')
    
    # 应用颜色标记到数据区域
    for row in range(2, 6):  # 第2到5行
        for col in range(2, 5):  # 第B到D列
            cell = ws.cell(row=row, column=col)
            cell_value = cell.value
            
            if cell_value is not None:
                try:
                    numeric_value = float(cell_value)
                    # 解决浮点数精度问题
                    if abs(numeric_value - round(numeric_value)) < 1e-10:
                        int_value = int(round(numeric_value))
                        
                        if int_value in color_fills:
                            cell.fill = color_fills[int_value]
                            cell.font = Font(color='000000')  # 黑色字体
                            print(f"  单元格 {cell.coordinate} 值 {int_value} 已标记颜色")
                except (ValueError, TypeError):
                    continue
    
    # 保存文件
    test_file = "test_colored_excel.xlsx"
    wb.save(test_file)
    print(f"✅ 测试Excel文件已保存: {test_file}")
    
    # 验证颜色是否正确应用
    verify_excel_colors(test_file)
    
    return test_file

def verify_excel_colors(file_path):
    """验证Excel文件中的颜色标记"""
    print(f"\n验证Excel文件颜色标记: {os.path.basename(file_path)}")
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active
        
        if ws is None:
            print("❌ 无法获取工作表")
            return False
        
        print(f"📊 工作表名称: {ws.title}")
        
        # 定义期望的颜色规则
        expected_colors = {
            2: 'FF0000',    # 红色
            6: 'FF0000',    # 红色
            -2: '00FF00',   # 绿色
            -6: '00FF00',   # 绿色
            8: 'FFFF00',    # 黄色
            10: 'FFCCCC',   # 淡红色
            15: 'FFCCCC',   # 淡红色
            -8: 'CCFFCC',   # 淡绿色
            12: 'CCFFCC'    # 淡绿色
        }
        
        # 检查数据区域的颜色
        color_check_count = 0
        correct_color_count = 0
        
        for row in range(2, 6):  # 第2到5行
            for col in range(2, 5):  # 第B到D列
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value is not None:
                    try:
                        # 尝试将值转换为数字
                        numeric_value = float(cell_value)
                        
                        # 解决浮点数精度问题
                        if abs(numeric_value - round(numeric_value)) < 1e-10:
                            int_value = int(round(numeric_value))
                            
                            # 检查是否在期望的颜色规则中
                            if int_value in expected_colors:
                                color_check_count += 1
                                
                                # 获取单元格的实际颜色
                                actual_color = None
                                if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                                    color_obj = cell.fill.start_color
                                    if hasattr(color_obj, 'rgb') and color_obj.rgb:
                                        actual_color = str(color_obj.rgb).upper()
                                
                                expected_color = expected_colors[int_value]
                                
                                # 检查颜色是否匹配
                                if actual_color and expected_color in actual_color:
                                    correct_color_count += 1
                                    print(f"   ✅ 单元格 {cell.coordinate} 值 {int_value} 颜色正确: {actual_color}")
                                else:
                                    print(f"   ❌ 单元格 {cell.coordinate} 值 {int_value} 颜色错误:")
                                    print(f"      期望: {expected_color}")
                                    print(f"      实际: {actual_color}")
                                    
                    except (ValueError, TypeError):
                        continue
        
        print(f"\n📊 验证结果:")
        print(f"   检查单元格数: {color_check_count}")
        print(f"   颜色正确数: {correct_color_count}")
        print(f"   正确率: {correct_color_count/color_check_count*100:.1f}%" if color_check_count > 0 else "   正确率: 0%")
        
        if color_check_count > 0 and correct_color_count == color_check_count:
            print("🎉 所有颜色标记都正确!")
            return True
        else:
            print("⚠️  部分颜色标记不正确或未找到匹配的单元格")
            return False
            
    except Exception as e:
        print(f"❌ 验证过程中出错: {str(e)}")
        return False

def simulate_email_attachment(file_path):
    """模拟邮件附件处理过程"""
    print(f"\n模拟邮件附件处理过程: {os.path.basename(file_path)}")
    
    try:
        # 创建邮件消息对象
        msg = MIMEMultipart()
        
        # 读取文件并创建附件
        filename = os.path.basename(file_path)
        
        # 根据文件扩展名自动检测MIME类型
        if filename.endswith('.xlsx'):
            main_type = 'application'
            sub_type = 'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif filename.endswith('.xls'):
            main_type = 'application'
            sub_type = 'vnd.ms-excel'
        else:
            main_type = 'application'
            sub_type = 'octet-stream'
        
        with open(file_path, 'rb') as f:
            part = MIMEBase(main_type, sub_type)
            part.set_payload(f.read())
            encoders.encode_base64(part)
            
            # 使用RFC2047编码的文件名
            encoded_filename = Header(filename, 'utf-8').encode()
            part.add_header(
                'Content-Disposition',
                f'attachment',
                filename=('utf-8', '', filename)
            )
            
            # 添加Content-Type头
            part.add_header('Content-Type', f'{main_type}/{sub_type}', name=('utf-8', '', filename))
            
            msg.attach(part)
        
        print("✅ 邮件附件处理完成")
        
        # 保存模拟的邮件附件到新文件
        attachment_file = "test_email_attachment.xlsx"
        with open(attachment_file, 'wb') as f:
            f.write(part.get_payload(decode=True))
        
        print(f"💾 模拟邮件附件已保存: {attachment_file}")
        
        # 验证附件中的颜色
        print("\n验证邮件附件中的颜色标记:")
        verify_excel_colors(attachment_file)
        
        return attachment_file
        
    except Exception as e:
        print(f"❌ 邮件附件处理过程中出错: {str(e)}")
        return None

def main():
    """主函数"""
    print("=== Excel颜色标记邮件附件测试 ===\n")
    
    # 1. 创建带颜色标记的测试Excel文件
    test_file = create_test_excel_with_colors()
    
    # 2. 模拟邮件附件处理过程
    attachment_file = simulate_email_attachment(test_file)
    
    # 3. 清理测试文件
    if os.path.exists(test_file):
        os.remove(test_file)
        print(f"\n🧹 已清理测试文件: {test_file}")
    
    if attachment_file and os.path.exists(attachment_file):
        # 保留附件文件供进一步分析
        print(f"📁 保留邮件附件文件: {attachment_file}")
        print(f"   请手动检查此文件的颜色标记是否正确")
    
    print("\n🎉 测试完成!")

if __name__ == "__main__":
    main()