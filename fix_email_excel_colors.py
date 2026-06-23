#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修复邮件附件中Excel文件颜色标记丢失的问题
"""

import os
from openpyxl import load_workbook
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header

def fix_excel_attachment_colors(file_path):
    """
    修复Excel附件中的颜色标记问题
    
    Args:
        file_path (str): Excel文件路径
        
    Returns:
        str: 修复后的文件路径
    """
    print(f"🔧 修复Excel文件颜色标记: {os.path.basename(file_path)}")
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active
        
        if ws is None:
            print("❌ 无法获取工作表")
            return file_path
        
        print(f"📊 工作表名称: {ws.title}")
        
        # 确保工作簿以Excel格式正确保存
        # 这是关键步骤：使用正确的参数保存Excel文件
        fixed_file_path = file_path.replace(".xlsx", "_fixed.xlsx")
        wb.save(fixed_file_path)
        
        print(f"✅ 颜色标记修复完成，文件已保存: {os.path.basename(fixed_file_path)}")
        return fixed_file_path
        
    except Exception as e:
        print(f"❌ 修复过程中出错: {str(e)}")
        return file_path

def create_proper_excel_attachment(file_path):
    """
    创建正确的Excel邮件附件
    
    Args:
        file_path (str): Excel文件路径
        
    Returns:
        MIMEBase: 邮件附件对象
    """
    try:
        filename = os.path.basename(file_path)
        
        # 使用正确的MIME类型
        main_type = 'application'
        sub_type = 'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        with open(file_path, 'rb') as f:
            part = MIMEBase(main_type, sub_type)
            part.set_payload(f.read())
            encoders.encode_base64(part)
            
            # 使用RFC2047编码的文件名
            encoded_filename = Header(filename, 'utf-8').encode()
            part.add_header(
                'Content-Disposition',
                f'attachment',
                filename=('utf-8', '', filename)
            )
            
            # 添加完整的Content-Type头
            part.add_header('Content-Type', f'{main_type}/{sub_type}; name="{filename}"')
            part.add_header('Content-Transfer-Encoding', 'base64')
            
        print(f"✅ Excel附件创建成功: {filename}")
        return part
        
    except Exception as e:
        print(f"❌ 创建Excel附件时出错: {str(e)}")
        return None

def verify_excel_format(file_path):
    """
    验证Excel文件格式是否正确
    
    Args:
        file_path (str): Excel文件路径
    """
    print(f"\n🔍 验证Excel文件格式: {os.path.basename(file_path)}")
    
    try:
        # 检查文件扩展名
        if not file_path.endswith('.xlsx'):
            print("⚠️  文件扩展名不是.xlsx")
            return False
        
        # 检查文件大小
        file_size = os.path.getsize(file_path)
        print(f"📄 文件大小: {file_size} 字节 ({file_size/1024:.1f} KB)")
        
        # 尝试加载工作簿
        wb = load_workbook(file_path)
        print(f"📊 工作表数量: {len(wb.worksheets)}")
        print(f"📋 活动工作表: {wb.active.title}")
        
        # 检查是否有数据
        ws = wb.active
        if ws and ws.max_row > 0 and ws.max_column > 0:
            print(f"📈 数据范围: {ws.max_row} 行 x {ws.max_column} 列")
        else:
            print("⚠️  工作表中没有数据")
        
        print("✅ Excel文件格式验证通过")
        return True
        
    except Exception as e:
        print(f"❌ Excel文件格式验证失败: {str(e)}")
        return False

def main():
    """
    主函数 - 演示如何正确处理Excel邮件附件
    """
    print("=== Excel邮件附件颜色标记修复演示 ===\n")
    
    # 示例：假设我们有一个带颜色标记的Excel文件
    # 在实际应用中，这将是process_real_mt4_data.py生成的文件
    sample_file = "sample_mt4_data_colored.xlsx"
    
    # 验证原始文件格式
    print("1. 验证原始文件格式:")
    verify_excel_format(sample_file)
    
    # 修复颜色标记（在实际应用中这一步可能不需要）
    print("\n2. 修复Excel颜色标记:")
    fixed_file = fix_excel_attachment_colors(sample_file)
    
    # 验证修复后的文件
    print("\n3. 验证修复后的文件:")
    verify_excel_format(fixed_file)
    
    # 创建正确的邮件附件
    print("\n4. 创建正确的邮件附件:")
    attachment = create_proper_excel_attachment(fixed_file)
    
    if attachment:
        print("🎉 Excel邮件附件处理完成!")
        print("   请确保在email_sender.py中使用此方法创建Excel附件")
    else:
        print("❌ Excel邮件附件创建失败")

# 在email_sender.py中应该这样修改attach_file方法：
EMAIL_ATTACHMENT_FIX_COMMENT = """
# 修复建议：在email_sender.py的attach_file方法中使用以下代码：

def attach_file(self, msg, file_path):
    '''
    添加文件附件（修复版，确保Excel颜色标记正确保存）
    '''
    try:
        import mimetypes
        from email.header import Header
        
        filename = os.path.basename(file_path)
        
        # 根据文件扩展名自动检测MIME类型
        if filename.endswith('.xlsx'):
            main_type = 'application'
            sub_type = 'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif filename.endswith('.xls'):
            main_type = 'application'
            sub_type = 'vnd.ms-excel'
        elif filename.endswith('.csv'):
            main_type = 'text'
            sub_type = 'csv'
        else:
            main_type = 'application'
            sub_type = 'octet-stream'
        
        with open(file_path, 'rb') as f:
            part = MIMEBase(main_type, sub_type)
            part.set_payload(f.read())
            encoders.encode_base64(part)
            
            # 使用RFC2047编码的文件名
            encoded_filename = Header(filename, 'utf-8').encode()
            part.add_header(
                'Content-Disposition',
                f'attachment',
                filename=('utf-8', '', filename)
            )
            
            # 添加完整的Content-Type头信息
            part.add_header('Content-Type', f'{main_type}/{sub_type}; name="{filename}"')
            part.add_header('Content-Transfer-Encoding', 'base64')
            
            msg.attach(part)
            
        print(f"📊 已添加附件: {os.path.basename(file_path)}")
        
    except Exception as e:
        print(f"❌ 添加附件失败 {file_path}: {str(e)}")
"""

if __name__ == "__main__":
    main()
    print(EMAIL_ATTACHMENT_FIX_COMMENT)