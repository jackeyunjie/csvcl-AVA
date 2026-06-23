#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证最新生成的Excel文件中的颜色标记
"""

import os
import openpyxl
import glob
from datetime import datetime

def get_latest_file(pattern):
    """获取最新的文件"""
    files = glob.glob(pattern)
    if not files:
        return None
    
    # 按文件的创建时间排序
    latest_file = max(files, key=os.path.getctime)
    return latest_file

def verify_color_rules(excel_file):
    """验证Excel文件中的颜色规则"""
    print(f"验证文件: {excel_file}")
    
    # 加载Excel文件
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 颜色规则映射
    color_rules = {
        2: 'FF0000',    # 红色背景
        3: 'FF0000',    # 红色背景
        4: 'FF0000',    # 红色背景
        5: 'FF0000',    # 红色背景
        6: 'FF0000',    # 红色背景
        7: 'FF0000',    # 红色背景
        -2: '00FF00',   # 绿色背景
        -3: '00FF00',   # 绿色背景
        -4: '00FF00',   # 绿色背景
        -5: '00FF00',   # 绿色背景
        -6: '00FF00',   # 绿色背景
        -7: '00FF00',   # 绿色背景
        8: 'FFFF00',    # 黄色背景
        10: 'FFCCCC',   # 淡红色背景
        11: 'FFCCCC',   # 淡红色背景
        12: 'FFCCCC',   # 淡红色背景
        13: 'FFCCCC',   # 淡红色背景
        14: 'FFCCCC',   # 淡红色背景
        15: 'FFCCCC',   # 淡红色背景
        -10: 'CCFFCC',  # 淡绿色背景
        -11: 'CCFFCC',  # 淡绿色背景
        -12: 'CCFFCC',  # 淡绿色背景
        -13: 'CCFFCC',  # 淡绿色背景
        -14: 'CCFFCC',  # 淡绿色背景
        -15: 'CCFFCC'   # 淡绿色背景
    }
    
    # 统计颜色单元格
    color_counts = {color: 0 for color in set(color_rules.values())}
    total_cells = 0
    
    # 查找特定区域E2:G43内的颜色单元格
    target_range = "E2:G43"
    start_col, start_row = 5, 2  # E2
    end_col, end_row = 7, 43    # G43
    
    # 检查每个单元格
    for row in range(start_row, min(end_row + 1, ws.max_row + 1)):
        for col in range(start_col, min(end_col + 1, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            cell_value = cell.value
            
            # 跳过空单元格
            if cell_value is None or str(cell_value).strip() == "":
                continue
            
            # 尝试转换为数字
            try:
                numeric_value = int(float(cell_value))
                total_cells += 1
                
                # 检查是否应该有颜色标记
                if numeric_value in color_rules:
                    expected_color = color_rules[numeric_value]
                    
                    # 获取实际的颜色标记
                    actual_color = None
                    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                        actual_color = cell.fill.start_color.rgb
                    
                    # 检查颜色是否匹配
                    color_match = False
                    if actual_color:
                        # 处理openpyxl的颜色格式（可能带有前缀）
                        if actual_color.endswith(expected_color):
                            color_match = True
                            color_counts[expected_color] += 1
                        # 直接检查颜色代码
                        elif actual_color == expected_color:
                            color_match = True
                            color_counts[expected_color] += 1
                    
                    # 显示结果
                    col_letter = get_column_letter(col)
                    if color_match:
                        print(f"✅ 单元格 {col_letter}{row} 值={numeric_value} 颜色正确: {actual_color}")
                    else:
                        print(f"❌ 单元格 {col_letter}{row} 值={numeric_value} 颜色错误: 期望={expected_color}, 实际={actual_color}")
                else:
                    col_letter = get_column_letter(col)
                    print(f"ℹ️  单元格 {col_letter}{row} 值={numeric_value} 不需要颜色标记")
            except (ValueError, TypeError):
                pass
    
    # 总结颜色统计
    print("\n颜色统计:")
    print(f"总检查单元格: {total_cells}")
    print(f"红色(FF0000): {color_counts.get('FF0000', 0)} 个单元格")
    print(f"绿色(00FF00): {color_counts.get('00FF00', 0)} 个单元格")
    print(f"黄色(FFFF00): {color_counts.get('FFFF00', 0)} 个单元格")
    print(f"淡红色(FFCCCC): {color_counts.get('FFCCCC', 0)} 个单元格")
    print(f"淡绿色(CCFFCC): {color_counts.get('CCFFCC', 0)} 个单元格")
    
    total_colored = sum(color_counts.values())
    print(f"总颜色标记: {total_colored} 个单元格")
    
    return total_colored > 0

def get_column_letter(col_num):
    """将列数字转换为字母（1=A, 2=B, ...）"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(col_num % 26 + ord('A')) + result
        col_num //= 26
    return result

if __name__ == "__main__":
    # 查找最新生成的Excel文件
    latest_excel = get_latest_file("*_MT4_colored.xlsx")
    
    if latest_excel:
        print(f"找到最新生成的Excel文件: {latest_excel}")
        if verify_color_rules(latest_excel):
            print("\n✅ 颜色规则验证通过！")
        else:
            print("\n❌ 颜色规则验证失败！")
    else:
        print("未找到_MT4_colored.xlsx文件")