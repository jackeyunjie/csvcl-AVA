#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV文件颜色标记程序
功能：
1. 从指定文件夹查找包含特定字符且在10分钟内创建的CSV文件
2. 对符合条件的数值进行颜色标记
"""

import os
import time
import glob
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import sys
from PIL import Image, ImageDraw, ImageFont
import io

class CSVColorMarker:
    def __init__(self, search_path=None, target_string="KVBt_@_D1", time_limit_minutes=5):
        """
        初始化CSV颜色标记器
        
        Args:
            search_path (str): 搜索路径
            target_string (str): 目标字符串，文件名需包含此字符串
            time_limit_minutes (int): 时间限制（分钟），默认7分钟
        """
        # 默认路径为MT4的Files目录
        if search_path is None:
            self.search_path = r"C:\Users\MECHREVO\AppData\Roaming\MetaQuotes\Terminal\50D8083188871EAB17316B22F188CFF7\MQL4\Files"
        else:
            self.search_path = search_path
            
        self.target_string = target_string
        self.time_limit_minutes = time_limit_minutes
        
        # 定义颜色映射
        self.color_rules = {
            2: 'FF0000',    # 红色背景
            3: 'FF0000',    # 红色背景
            4: 'FF0000',    # 红色背景
            5: 'FF0000',    # 红色背景
            6: 'FF0000',    # 红色背景
            7: 'FF0000',    # 红色背景
            10: 'FFCCCC',   # 淡红色背景
            11: 'FFCCCC',   # 淡红色背景
            12: 'FFCCCC',   # 淡红色背景
            13: 'FFCCCC',   # 淡红色背景
            14: 'FFCCCC',   # 淡红色背景
            15: 'FFCCCC',   # 淡红色背景
            -2: '00FF00',   # 绿色背景
            -3: '00FF00',   # 绿色背景
            -4: '00FF00',   # 绿色背景
            -5: '00FF00',   # 绿色背景
            -6: '00FF00',   # 绿色背景
            -7: '00FF00',   # 绿色背景
            -10: 'CCFFCC',  # 淡绿色背景
            -11: 'CCFFCC',  # 淡绿色背景
            -12: 'CCFFCC',  # 淡绿色背景
            -13: 'CCFFCC',  # 淡绿色背景
            -14: 'CCFFCC',  # 淡绿色背景
            -15: 'CCFFCC',  # 淡绿色背景
            8: 'FFFF00'     # 黄色背景
        }
    
    def find_recent_csv_files(self):
        """
        查找符合条件的CSV文件
        
        Returns:
            list: 符合条件的CSV文件路径列表
        """
        print(f"正在搜索目录: {self.search_path}")
        print(f"查找包含字符串: '{self.target_string}' 的文件")
        print(f"时间限制: {self.time_limit_minutes} 分钟内创建的文件")
        
        if not os.path.exists(self.search_path):
            print(f"错误：目录不存在 - {self.search_path}")
            return []
        
        current_time = datetime.now()
        time_threshold = current_time - timedelta(minutes=self.time_limit_minutes)
        
        # 搜索CSV文件
        csv_pattern = os.path.join(self.search_path, "*.csv")
        all_csv_files = glob.glob(csv_pattern)
        
        print(f"找到 {len(all_csv_files)} 个CSV文件")
        
        matching_files = []
        
        for file_path in all_csv_files:
            filename = os.path.basename(file_path)
            
            # 检查文件名是否包含目标字符串
            if self.target_string in filename:
                # 检查文件创建时间
                creation_time = datetime.fromtimestamp(os.path.getctime(file_path))
                
                if creation_time >= time_threshold:
                    matching_files.append(file_path)
                    print(f"[OK] 找到匹配文件: {filename}")
                    print(f"  创建时间: {creation_time.strftime('%Y-%m-%d %H:%M:%S')}")
                else:
                    print(f"[X] 文件太旧: {filename} (创建于 {creation_time.strftime('%Y-%m-%d %H:%M:%S')})")
            else:
                print(f"[X] 文件名不匹配: {filename}")
        
        print(f"\n总共找到 {len(matching_files)} 个符合条件的文件")
        return matching_files
    
    def apply_color_formatting(self, df, worksheet, target_range="E2:G40"):
        """
        对指定单元格区域应用颜色格式
        
        Args:
            df (pandas.DataFrame): 数据框
            worksheet: openpyxl工作表对象
            target_range (str): 目标单元格区域，默认为"E2:G40"
        """
        # 解析目标区域
        start_col, start_row, end_col, end_row = self.parse_range(target_range)
        
        print(f"正在对区域 {target_range} 进行颜色标记...")
        print(f"区域范围: 列{start_col}-{end_col}, 行{start_row}-{end_row}")
        
        # 创建颜色填充和字体对象
        color_fills = {}
        fonts = {}
        for value, color_code in self.color_rules.items():
            color_fills[value] = PatternFill(start_color=color_code, 
                                           end_color=color_code, 
                                           fill_type='solid')
            # 为红色背景的单元格（数值2、4、6）设置黑色字体
            if color_code == 'FF0000':  # 红色背景
                fonts[value] = Font(color='000000')  # 黑色字体
            # 为淡红色背景的单元格（数值10-15）设置黑色字体
            elif color_code == 'FFCCCC':  # 淡红色背景
                fonts[value] = Font(color='000000')  # 黑色字体
            # 为绿色背景的单元格（数值-2、-4、-6）设置黑色字体
            elif color_code == '00FF00':  # 绿色背景
                fonts[value] = Font(color='000000')  # 黑色字体
            # 为淡绿色背景的单元格（数值-10到-15）设置黑色字体
            elif color_code == 'CCFFCC':  # 淡绿色背景
                fonts[value] = Font(color='000000')  # 黑色字体
            else:
                fonts[value] = Font(color='000000')  # 黑色字体（默认）
        
        colored_count = 0
        
        # 遍历指定区域的单元格
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                try:
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value
                    
                    if cell_value is not None:
                        # 尝试将值转换为数字
                        numeric_value = float(cell_value)
                        
                        # 解决浮点数精度问题：检查是否接近整数值
                        if abs(numeric_value - round(numeric_value)) < 1e-10:
                            int_value = int(round(numeric_value))
                            
                            # 检查整数值是否在颜色规则中
                            if int_value in self.color_rules:
                                cell.fill = color_fills[int_value]
                                cell.font = fonts[int_value]  # 设置字体颜色
                                col_letter = self.get_column_letter(col)
                                # 根据实际的颜色规则判断字体颜色
                                font_color = "黑色"
                                print(f"单元格 {col_letter}{row} 值 {int_value} 已标记为相应颜色，字体颜色: {font_color}")
                                colored_count += 1
                                
                except (ValueError, TypeError):
                    # 如果不是数字，跳过
                    continue
        
        print(f"[OK] 共标记了 {colored_count} 个单元格")
    
    def parse_range(self, range_str):
        """
        解析Excel区域字符串（如"E2:G40"）
        
        Args:
            range_str (str): 区域字符串
            
        Returns:
            tuple: (start_col, start_row, end_col, end_row)
        """
        try:
            start_cell, end_cell = range_str.split(':')
            
            # 解析起始单元格
            start_col = self.column_letter_to_number(start_cell.strip('0123456789'))
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            
            # 解析结束单元格
            end_col = self.column_letter_to_number(end_cell.strip('0123456789'))
            end_row = int(''.join(filter(str.isdigit, end_cell)))
            
            return start_col, start_row, end_col, end_row
            
        except Exception as e:
            print(f"解析区域时出错: {e}, 使用默认区域")
            return 5, 2, 7, 40  # E2:G40的默认值
    
    def column_letter_to_number(self, letter):
        """
        将列字母转换为数字（A=1, B=2, ...）
        
        Args:
            letter (str): 列字母
            
        Returns:
            int: 列数字
        """
        result = 0
        for char in letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def get_column_letter(self, col_num):
        """
        将列数字转换为字母（1=A, 2=B, ...）
        
        Args:
            col_num (int): 列数字
            
        Returns:
            str: 列字母
        """
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def adjust_column_width(self, worksheet):
        """
        调整列宽，将B列的宽度设置为其他列的1.5倍
        
        Args:
            worksheet: openpyxl工作表对象
        """
        try:
            # 设置默认列宽
            default_width = 12
            
            # 为所有列设置默认宽度
            for col in range(1, worksheet.max_column + 1):
                col_letter = self.get_column_letter(col)
                worksheet.column_dimensions[col_letter].width = default_width
            
            # 将B列的宽度设置为默认宽度的1.5倍
            worksheet.column_dimensions['B'].width = default_width * 1.5
            
            print(f"[OK] 已调整B列宽度为其他列的1.5倍（B列: {default_width * 1.5}, 其他列: {default_width}）")
            
        except Exception as e:
            print(f"[ERROR] 调整列宽时出错: {str(e)}")
    
    def save_range_as_image(self, worksheet, range_str="A1:M40", output_path=None, include_headers=True):
        """
        将指定区域保存为图片，包含行号和列号
        
        Args:
            worksheet: openpyxl工作表对象
            range_str (str): 要保存的区域，默认为"A1:H40"
            output_path (str): 输出图片路径，如果为None则自动生成
            include_headers (bool): 是否包含行号和列号，默认为True
        
        Returns:
            str: 保存的图片路径
        """
        try:
            print(f"正在将区域 {range_str} 保存为图片（包含行列号）...")
            
            # 解析区域
            start_col, start_row, end_col, end_row = self.parse_range(range_str)
            
            # 设置单元格和图片的基本尺寸
            cell_width = 80
            cell_height = 25
            header_width = 40  # 行号列宽度
            header_height = 25  # 列号行高度
            
            # 计算图片尺寸（包含行号和列号区域）
            if include_headers:
                img_width = header_width + (end_col - start_col + 1) * cell_width
                img_height = header_height + (end_row - start_row + 1) * cell_height
            else:
                img_width = (end_col - start_col + 1) * cell_width
                img_height = (end_row - start_row + 1) * cell_height
            
            # 创建图片
            img = Image.new('RGB', (img_width, img_height), 'white')
            draw = ImageDraw.Draw(img)
            
            # 尝试加载字体
            try:
                font = ImageFont.truetype("arial.ttf", 10)
                header_font = ImageFont.truetype("arial.ttf", 9)
            except:
                try:
                    font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 10)
                    header_font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 9)
                except:
                    font = ImageFont.load_default()
                    header_font = ImageFont.load_default()
            
            # 定义颜色映射（PIL使用RGB格式）
            color_map = {
                '00FF0000': (255, 0, 0),    # 红色背景（openpyxl格式）
                'FF0000': (255, 0, 0),      # 红色背景（标准格式）
                '0000FF00': (0, 255, 0),    # 绿色背景（openpyxl格式）
                '00FF00': (0, 255, 0),      # 绿色背景（标准格式）
                '00FFFF00': (255, 255, 0),  # 黄色背景（openpyxl格式）
                'FFFF00': (255, 255, 0),    # 黄色背景（标准格式）
                '00FFCCCC': (255, 204, 204), # 淡红色背景（openpyxl格式）
                'FFCCCC': (255, 204, 204),   # 淡红色背景（标准格式）
                '00CCFFCC': (204, 255, 204), # 淡绿色背景（openpyxl格式）
                'CCFFCC': (204, 255, 204),   # 淡绿色背景（标准格式）
                'FFFFFF': (255, 255, 255),  # 白色背景（默认）
                '00FFFFFF': (255, 255, 255), # 白色背景（openpyxl格式）
                '00000000': (255, 255, 255)  # 透明色作为白色处理
            }
            
            # 绘制行号和列号的偏移量
            offset_x = header_width if include_headers else 0
            offset_y = header_height if include_headers else 0
            
            # 绘制列号标题（如果包含标题）
            if include_headers:
                # 绘制左上角空白区域
                draw.rectangle([0, 0, header_width, header_height], 
                             fill=(220, 220, 220), outline=(0, 0, 0))
                
                # 绘制列号
                for col in range(start_col, end_col + 1):
                    x = offset_x + (col - start_col) * cell_width
                    y = 0
                    
                    # 绘制列号背景
                    draw.rectangle([x, y, x + cell_width, y + header_height], 
                                 fill=(220, 220, 220), outline=(0, 0, 0))
                    
                    # 绘制列号文本
                    col_letter = self.get_column_letter(col)
                    bbox = draw.textbbox((0, 0), col_letter, font=header_font)
                    text_width = bbox[2] - bbox[0]
                    text_height = bbox[3] - bbox[1]
                    text_x = x + (cell_width - text_width) // 2
                    text_y = y + (header_height - text_height) // 2
                    draw.text((text_x, text_y), col_letter, fill=(0, 0, 0), font=header_font)
                
                # 绘制行号
                for row in range(start_row, end_row + 1):
                    x = 0
                    y = offset_y + (row - start_row) * cell_height
                    
                    # 绘制行号背景
                    draw.rectangle([x, y, x + header_width, y + cell_height], 
                                 fill=(220, 220, 220), outline=(0, 0, 0))
                    
                    # 绘制行号文本
                    row_text = str(row)
                    bbox = draw.textbbox((0, 0), row_text, font=header_font)
                    text_width = bbox[2] - bbox[0]
                    text_height = bbox[3] - bbox[1]
                    text_x = x + (header_width - text_width) // 2
                    text_y = y + (cell_height - text_height) // 2
                    draw.text((text_x, text_y), row_text, fill=(0, 0, 0), font=header_font)
            
            # 绘制数据单元格
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    # 计算单元格在图片中的位置
                    x = offset_x + (col - start_col) * cell_width
                    y = offset_y + (row - start_row) * cell_height
                    
                    # 获取单元格
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value if cell.value is not None else ""
                    
                    # 确定背景颜色
                    bg_color = (255, 255, 255)  # 默认白色
                    
                    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                        # 获取颜色代码
                        color_obj = cell.fill.start_color
                        color_code = None
                        
                        # 尝试获取RGB值
                        if hasattr(color_obj, 'rgb') and color_obj.rgb:
                            color_code = str(color_obj.rgb)
                        
                        # 如果有颜色代码且不是透明或默认值
                        if color_code and color_code in color_map:
                            bg_color = color_map[color_code]
                    
                    # 绘制单元格背景
                    draw.rectangle([x, y, x + cell_width, y + cell_height], 
                                 fill=bg_color, outline=(0, 0, 0))
                    
                    # 绘制单元格文本
                    if cell_value is not None and str(cell_value).strip() != "":
                        text = str(cell_value)
                        # 计算文本位置（居中）
                        bbox = draw.textbbox((0, 0), text, font=font)
                        text_width = bbox[2] - bbox[0]
                        text_height = bbox[3] - bbox[1]
                        text_x = x + (cell_width - text_width) // 2
                        text_y = y + (cell_height - text_height) // 2
                        
                        # 绘制文本（黑色字体）
                        draw.text((text_x, text_y), text, fill=(0, 0, 0), font=font)
            
            # 生成输出路径
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = os.path.join(os.path.dirname(__file__), f"excel_full_table_{timestamp}.png")
            
            # 保存图片
            img.save(output_path, 'PNG')
            print(f"[OK] 已保存完整表格图片: {os.path.basename(output_path)}")
            print(f"  完整路径: {os.path.abspath(output_path)}")
            print(f"  图片尺寸: {img_width} x {img_height} 像素")
            print(f"  包含行列号: {'是' if include_headers else '否'}")
            
            return output_path
            
        except Exception as e:
            print(f"[ERROR] 保存图片时出错: {str(e)}")
            return None
    
    def process_csv_file(self, file_path):
        """
        处理单个CSV文件
        
        Args:
            file_path (str): CSV文件路径
        """
        try:
            print(f"\n正在处理文件: {os.path.basename(file_path)}")
            
            # 生成输出文件名（提前定义base_name）
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            
            # 读取CSV文件
            df = pd.read_csv(file_path, encoding='utf-8')
            print(f"文件包含 {len(df)} 行，{len(df.columns)} 列")
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Colored_Data"
            
            # 将数据框写入工作表
            if ws is not None:
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
            
            # 应用颜色格式到指定区域（E2:G40）
            if ws is not None:
                self.apply_color_formatting(df, ws, "E2:G40")
                # 调整B列的列宽为其他列的1.5倍
                self.adjust_column_width(ws)
                
                # 保存A1:M40区域为图片（包含行列号）
                image_filename = f"{base_name}_full_table.png"
                image_path = os.path.join(os.path.dirname(__file__), image_filename)
                self.save_range_as_image(ws, "A1:M40", image_path, include_headers=True)
            
            # 保存Excel文件
            output_filename = f"{base_name}_colored.xlsx"
            output_path = os.path.join(os.path.dirname(__file__), output_filename)
            
            # 保存文件
            wb.save(output_path)
            print(f"[OK] 已保存彩色标记文件: {output_filename}")
            print(f"  完整路径: {os.path.abspath(output_path)}")
            
            return output_path
            
        except Exception as e:
            print(f"[ERROR] 处理文件时出错: {str(e)}")
            return None
    
    def run(self):
        """
        运行主程序
        """
        print("=== CSV文件颜色标记程序 ===")
        print(f"颜色规则:")
        print(f"  红色背景+黑色字体: 2, 3, 4, 5, 6, 7")
        print(f"  淡红色背景+黑色字体: 10, 11, 12, 13, 14, 15")
        print(f"  绿色背景+黑色字体: -2, -3, -4, -5, -6, -7")
        print(f"  淡绿色背景+黑色字体: -10, -11, -12, -13, -14, -15")
        print(f"  黄色背景+黑色字体: 8")
        print("=" * 40)
        
        # 查找符合条件的文件
        matching_files = self.find_recent_csv_files()
        
        if not matching_files:
            print("\n没有找到符合条件的文件。")
            return
        
        # 处理每个文件
        processed_files = []
        for file_path in matching_files:
            result = self.process_csv_file(file_path)
            if result:
                processed_files.append(result)
        
        print(f"\n=== 处理完成 ===")
        print(f"成功处理 {len(processed_files)} 个文件")
        for file_path in processed_files:
            print(f"  - {os.path.basename(file_path)}")
            print(f"    完整路径: {os.path.abspath(file_path)}")

def main():
    """
    主函数
    """
    parser = argparse.ArgumentParser(description='CSV文件颜色标记程序')
    parser.add_argument('--path', type=str, help='搜索路径')
    parser.add_argument('--string', type=str, default='KVBt_@_D1', help='目标字符串')
    parser.add_argument('--minutes', type=int, default=5, help='时间限制（分钟）')
    
    args = parser.parse_args()
    
    # 创建并运行颜色标记器
    marker = CSVColorMarker(
        search_path=args.path,
        target_string=args.string,
        time_limit_minutes=args.minutes
    )
    
    marker.run()

if __name__ == "__main__":
    main()