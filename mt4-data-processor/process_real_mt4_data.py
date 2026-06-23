#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
处理真实MT4数据的完整流程
不使用模拟数据，直接处理MT4生成的真实CSV文件
"""

import os
import time
import glob
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image, ImageDraw, ImageFont
from email_sender import EmailSender

class RealMT4DataProcessor:
    def __init__(self, mt4_path=None, target_string="KVBt_@_D1", time_limit_minutes=10, enable_email=False, recipients=None):
        """
        初始化真实MT4数据处理器
        
        Args:
            mt4_path (str): MT4数据路径
            target_string (str): 目标字符串，文件名需包含此字符串
            time_limit_minutes (int): 时间限制（分钟），默认10分钟
            enable_email (bool): 是否启用邮件发送
            recipients (list): 收件人邮箱列表
        """
        # 默认MT4路径
        if mt4_path is None:
            self.mt4_path = r"C:\Users\MECHREVO\AppData\Roaming\MetaQuotes\Terminal\6E3DD078C0E9FF56930185F9E6CDDE71\MQL4\Files"
        else:
            self.mt4_path = mt4_path
            
        self.target_string = target_string
        self.time_limit_minutes = time_limit_minutes
        self.enable_email = enable_email
        # 默认收件人配置：447372703@qq.com 和 1300893414@qq.com
        if recipients is None:
            self.recipients = ["447372703@qq.com", "1300893414@qq.com"]
        else:
            self.recipients = recipients
        
        # 初始化邮件发送器
        if self.enable_email:
            try:
                self.email_sender = EmailSender()
                print(f"✅ 邮件发送功能已启用")
                print(f"📧 收件人: {', '.join(self.recipients)}")
            except Exception as e:
                print(f"⚠️  邮件发送器初始化失败: {str(e)}")
                self.enable_email = False
        
        # 定义颜色映射
        self.color_rules = {
            2: 'FF0000',    # 红色背景
            4: 'FF0000',    # 红色背景
            6: 'FF0000',    # 红色背景
            -2: '00FF00',   # 绿色背景
            -4: '00FF00',   # 绿色背景
            -6: '00FF00',   # 绿色背景
            8: 'FFFF00'     # 黄色背景
        }
    
    def find_real_mt4_files(self):
        """
        查找真实的MT4 CSV文件
        
        Returns:
            list: 符合条件的CSV文件路径列表
        """
        print(f"=== 查找真实MT4数据文件 ===")
        print(f"搜索路径: {self.mt4_path}")
        print(f"目标字符串: '{self.target_string}'")
        print(f"时间限制: {self.time_limit_minutes} 分钟内创建的文件")
        
        if not os.path.exists(self.mt4_path):
            print(f"❌ MT4目录不存在: {self.mt4_path}")
            print("请确认MT4安装路径或手动指定正确路径")
            return []
        
        current_time = datetime.now()
        time_threshold = current_time - timedelta(minutes=self.time_limit_minutes)
        
        # 搜索CSV文件
        csv_pattern = os.path.join(self.mt4_path, "*.csv")
        all_csv_files = glob.glob(csv_pattern)
        
        print(f"📁 在MT4目录找到 {len(all_csv_files)} 个CSV文件")
        
        matching_files = []
        
        for file_path in all_csv_files:
            filename = os.path.basename(file_path)
            
            # 检查文件名是否包含目标字符串
            if self.target_string in filename:
                # 检查文件创建时间
                creation_time = datetime.fromtimestamp(os.path.getctime(file_path))
                
                if creation_time >= time_threshold:
                    matching_files.append(file_path)
                    print(f"✅ 找到匹配的MT4文件: {filename}")
                    print(f"   创建时间: {creation_time.strftime('%Y-%m-%d %H:%M:%S')}")
                    print(f"   文件大小: {os.path.getsize(file_path)} 字节")
                else:
                    print(f"⏰ 文件创建时间超出限制: {filename}")
                    print(f"   创建时间: {creation_time.strftime('%Y-%m-%d %H:%M:%S')}")
            else:
                continue  # 不显示不匹配的文件，减少输出
        
        print(f"\n📊 总共找到 {len(matching_files)} 个符合条件的MT4文件")
        return matching_files
    
    def process_real_csv_data(self, file_path):
        """
        处理真实的CSV数据
        
        Args:
            file_path (str): CSV文件路径
        """
        try:
            print(f"\n=== 处理真实MT4数据 ===")
            print(f"📄 文件: {os.path.basename(file_path)}")
            
            # 生成输出文件名
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            
            # 第1步：读取真实CSV文件
            print(f"🔄 第1步：读取真实MT4 CSV数据...")
            df = pd.read_csv(file_path, encoding='utf-8')
            print(f"   ✅ 成功读取: {len(df)} 行，{len(df.columns)} 列")
            print(f"   📋 列名: {list(df.columns)}")
            
            # 显示数据预览
            print(f"   📊 数据预览:")
            print(df.head().to_string())
            
            # 数据变化分析
            data_changes = self.analyze_data_changes(df)
            
            # 第2步：创建Excel文件并应用颜色标记
            print(f"\n🔄 第2步：创建Excel文件并标记颜色...")
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "MT4_Data_Colored"
            
            # 将数据框写入工作表
            if ws is not None:
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
            
            # 应用颜色格式到指定区域（E2:G40）
            if ws is not None:
                colored_count = self.apply_color_formatting(df, ws, "E2:G40")
                print(f"   ✅ 颜色标记完成：共标记了 {colored_count} 个单元格")
                
                # 调整B列的列宽为其他列的1.5倍
                self.adjust_column_width(ws)
                print(f"   ✅ 列宽调整完成：B列宽度为其他列的1.5倍")
            
            # 保存Excel文件
            excel_filename = f"{base_name}_MT4_colored.xlsx"
            excel_path = os.path.join(os.path.dirname(__file__), excel_filename)
            wb.save(excel_path)
            print(f"   ✅ Excel文件已保存: {excel_filename}")
            
            # 第3步：生成A1:M40范围的截图
            print(f"\n🔄 第3步：生成A1:M40范围截图...")
            if ws is not None:
                image_filename = f"{base_name}_MT4_screenshot.jpg"  # 改为JPG格式
                image_path = os.path.join(os.path.dirname(__file__), image_filename)
                self.save_range_as_image(ws, "A1:M40", image_path, include_headers=True, image_format="JPG")
                print(f"   ✅ 截图已保存: {image_filename}")
            
            print(f"\n🎉 MT4数据处理完成！")
            print(f"📁 生成文件:")
            print(f"   📊 Excel: {os.path.abspath(excel_path)}")
            print(f"   🖼️  图片: {os.path.abspath(image_path)}")
            
            return excel_path, image_path, data_changes
            
        except Exception as e:
            print(f"❌ 处理MT4数据时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            return None, None, []
    
    def apply_color_formatting(self, df, worksheet, target_range="E2:G40"):
        """
        对指定单元格区域应用颜色格式
        """
        # 解析目标区域
        start_col, start_row, end_col, end_row = self.parse_range(target_range)
        
        # 创建颜色填充和字体对象
        color_fills = {}
        fonts = {}
        for value, color_code in self.color_rules.items():
            color_fills[value] = PatternFill(start_color=color_code, 
                                           end_color=color_code, 
                                           fill_type='solid')
            fonts[value] = Font(color='000000')  # 黑色字体
        
        colored_count = 0
        
        # 遍历指定区域的单元格
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                try:
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value
                    
                    if cell_value is not None:
                        # 尝试将值转换为数字
                        numeric_value = float(cell_value)
                        
                        # 检查是否需要着色
                        if numeric_value in self.color_rules:
                            cell.fill = color_fills[numeric_value]
                            cell.font = fonts[numeric_value]
                            colored_count += 1
                            
                except (ValueError, TypeError):
                    # 如果不是数字，跳过
                    continue
        
        return colored_count
    
    def parse_range(self, range_str):
        """解析Excel区域字符串"""
        try:
            start_cell, end_cell = range_str.split(':')
            
            # 解析起始单元格
            start_col = self.column_letter_to_number(start_cell.strip('0123456789'))
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            
            # 解析结束单元格
            end_col = self.column_letter_to_number(end_cell.strip('0123456789'))
            end_row = int(''.join(filter(str.isdigit, end_cell)))
            
            return start_col, start_row, end_col, end_row
            
        except Exception as e:
            print(f"解析区域时出错: {e}, 使用默认区域")
            return 5, 2, 7, 40  # E2:G40的默认值
    
    def column_letter_to_number(self, letter):
        """将列字母转换为数字"""
        result = 0
        for char in letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def get_column_letter(self, col_num):
        """将列数字转换为字母"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def adjust_column_width(self, worksheet):
        """调整列宽"""
        try:
            default_width = 12
            
            # 为所有列设置默认宽度
            for col in range(1, worksheet.max_column + 1):
                col_letter = self.get_column_letter(col)
                worksheet.column_dimensions[col_letter].width = default_width
            
            # 将B列的宽度设置为默认宽度的1.5倍
            worksheet.column_dimensions['B'].width = default_width * 1.5
            
        except Exception as e:
            print(f"调整列宽时出错: {str(e)}")
    
    def analyze_data_changes(self, df):
        """
        分析数据变化，特别关注重要数值的变化
        
        Args:
            df: 包含MT4数据的DataFrame
            
        Returns:
            list: 变化分析结果列表
        """
        try:
            print(f"\n🔍 === 数据变化分析 ===")
            
            # 重要数值列表
            important_values = [2, 6, -2, -6]
            
            # 需要分析的列（E、F、G对应MN1、W1、D1）
            analysis_columns = ['MN1', 'W1', 'D1']
            
            changes = []
            
            # 按品种分组（每3行一个品种）
            total_rows = len(df)
            
            for i in range(0, total_rows, 3):
                if i + 1 < total_rows:  # 确保至少有今天和昨天的数据
                    today_row = df.iloc[i]
                    yesterday_row = df.iloc[i + 1]
                    
                    symbol = today_row['SymbolName']
                    today_date = today_row['TIME']
                    yesterday_date = yesterday_row['TIME']
                    
                    # 分析每个重要列的变化
                    for col in analysis_columns:
                        today_value = today_row[col]
                        yesterday_value = yesterday_row[col]
                        
                        # 检查是否变成了重要数值
                        if (today_value in important_values and 
                            yesterday_value != today_value):
                            
                            change_info = {
                                'symbol': symbol,
                                'column': col,
                                'today_date': today_date,
                                'yesterday_date': yesterday_date,
                                'from_value': yesterday_value,
                                'to_value': today_value
                            }
                            
                            changes.append(change_info)
                            
                            print(f"📊 发现重要变化: {symbol} {col}列 {yesterday_value} → {today_value}")
            
            print(f"✅ 数据变化分析完成，发现 {len(changes)} 个重要变化")
            return changes
            
        except Exception as e:
            print(f"❌ 数据变化分析失败: {str(e)}")
            return []
    
    def save_range_as_image(self, worksheet, range_str="A1:M40", output_path=None, include_headers=True, image_format="JPG"):
        """将指定区域保存为图片，包含行号和列号"""
        try:
            print(f"   🖼️  正在生成 {range_str} 区域截图（{image_format}格式）...")
            
            # 解析区域
            start_col, start_row, end_col, end_row = self.parse_range(range_str)
            
            # 设置单元格和图片的基本尺寸
            cell_width = 80
            cell_height = 25
            header_width = 40
            header_height = 25
            
            # 计算图片尺寸
            if include_headers:
                img_width = header_width + (end_col - start_col + 1) * cell_width
                img_height = header_height + (end_row - start_row + 1) * cell_height
            else:
                img_width = (end_col - start_col + 1) * cell_width
                img_height = (end_row - start_row + 1) * cell_height
            
            # 创建图片
            img = Image.new('RGB', (img_width, img_height), 'white')
            draw = ImageDraw.Draw(img)
            
            # 尝试加载字体
            try:
                font = ImageFont.truetype("arial.ttf", 10)
                header_font = ImageFont.truetype("arial.ttf", 9)
            except:
                try:
                    font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 10)
                    header_font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 9)
                except:
                    font = ImageFont.load_default()
                    header_font = ImageFont.load_default()
            
            # 定义颜色映射
            color_map = {
                '00FF0000': (255, 0, 0),    # 红色背景
                'FF0000': (255, 0, 0),      # 红色背景
                '0000FF00': (0, 255, 0),    # 绿色背景
                '00FF00': (0, 255, 0),      # 绿色背景
                '00FFFF00': (255, 255, 0),  # 黄色背景
                'FFFF00': (255, 255, 0),    # 黄色背景
                'FFFFFF': (255, 255, 255),  # 白色背景
                '00FFFFFF': (255, 255, 255),
                '00000000': (255, 255, 255)
            }
            
            # 绘制偏移量
            offset_x = header_width if include_headers else 0
            offset_y = header_height if include_headers else 0
            
            # 绘制行号和列号
            if include_headers:
                # 左上角空白区域
                draw.rectangle([0, 0, header_width, header_height], 
                             fill=(220, 220, 220), outline=(0, 0, 0))
                
                # 绘制列号
                for col in range(start_col, end_col + 1):
                    x = offset_x + (col - start_col) * cell_width
                    y = 0
                    
                    draw.rectangle([x, y, x + cell_width, y + header_height], 
                                 fill=(220, 220, 220), outline=(0, 0, 0))
                    
                    col_letter = self.get_column_letter(col)
                    bbox = draw.textbbox((0, 0), col_letter, font=header_font)
                    text_width = bbox[2] - bbox[0]
                    text_height = bbox[3] - bbox[1]
                    text_x = x + (cell_width - text_width) // 2
                    text_y = y + (header_height - text_height) // 2
                    draw.text((text_x, text_y), col_letter, fill=(0, 0, 0), font=header_font)
                
                # 绘制行号
                for row in range(start_row, end_row + 1):
                    x = 0
                    y = offset_y + (row - start_row) * cell_height
                    
                    draw.rectangle([x, y, x + header_width, y + cell_height], 
                                 fill=(220, 220, 220), outline=(0, 0, 0))
                    
                    row_text = str(row)
                    bbox = draw.textbbox((0, 0), row_text, font=header_font)
                    text_width = bbox[2] - bbox[0]
                    text_height = bbox[3] - bbox[1]
                    text_x = x + (header_width - text_width) // 2
                    text_y = y + (cell_height - text_height) // 2
                    draw.text((text_x, text_y), row_text, fill=(0, 0, 0), font=header_font)
            
            # 绘制数据单元格
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    x = offset_x + (col - start_col) * cell_width
                    y = offset_y + (row - start_row) * cell_height
                    
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value if cell.value is not None else ""
                    
                    # 确定背景颜色
                    bg_color = (255, 255, 255)  # 默认白色
                    
                    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                        color_obj = cell.fill.start_color
                        color_code = None
                        
                        if hasattr(color_obj, 'rgb') and color_obj.rgb:
                            color_code = str(color_obj.rgb)
                        
                        if color_code and color_code in color_map:
                            bg_color = color_map[color_code]
                    
                    # 绘制单元格背景
                    draw.rectangle([x, y, x + cell_width, y + cell_height], 
                                 fill=bg_color, outline=(0, 0, 0))
                    
                    # 绘制单元格文本
                    if cell_value is not None and str(cell_value).strip() != "":
                        text = str(cell_value)
                        bbox = draw.textbbox((0, 0), text, font=font)
                        text_width = bbox[2] - bbox[0]
                        text_height = bbox[3] - bbox[1]
                        text_x = x + (cell_width - text_width) // 2
                        text_y = y + (cell_height - text_height) // 2
                        
                        draw.text((text_x, text_y), text, fill=(0, 0, 0), font=font)
            
            # 生成输出路径
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                ext = "jpg" if image_format.upper() == "JPG" else "png"
                output_path = os.path.join(os.path.dirname(__file__), f"MT4_screenshot_{timestamp}.{ext}")
            
            # 保存图片（支持JPG和PNG格式）
            if image_format.upper() == "JPG":
                # JPG格式需要RGB模式，不支持透明
                img.save(output_path, 'JPEG', quality=95, optimize=True)
                print(f"   ✅ JPG截图保存成功（质量: 95%）")
            else:
                # PNG格式支持透明
                img.save(output_path, 'PNG')
                print(f"   ✅ PNG截图保存成功")
            
            print(f"      📁 路径: {os.path.abspath(output_path)}")
            print(f"      📐 尺寸: {img_width} x {img_height} 像素")
            print(f"      💾 格式: {image_format.upper()}")
            print(f"      🏷️  包含边框: {'是' if include_headers else '否'}")
            
            return output_path
            
        except Exception as e:
            print(f"   ❌ 生成截图时出错: {str(e)}")
            return None
    
    def run(self):
        """
        运行完整的MT4数据处理流程
        """
        print("🚀 === 真实MT4数据处理程序启动 ===")
        print("📋 处理流程:")
        print("   1️⃣  查找并读取真实MT4 CSV文件")
        print("   2️⃣  生成Excel文件并应用颜色标记")
        print("   3️⃣  生成A1:M40范围的完整截图")
        if self.enable_email:
            print("   4️⃣  发送邮件报告")
        print("=" * 50)
        
        # 查找真实MT4文件
        mt4_files = self.find_real_mt4_files()
        
        if not mt4_files:
            print("\n❌ 未找到符合条件的MT4文件")
            print("💡 建议:")
            print("   - 检查MT4是否正在运行并生成数据")
            print("   - 确认文件名包含 'KVBt_@_D1'")
            print("   - 增加时间限制范围")
            return
        
        # 处理每个MT4文件
        success_count = 0
        excel_files = []
        image_files = []
        all_changes = []  # 收集所有文件的变化
        
        for file_path in mt4_files:
            excel_path, image_path, data_changes = self.process_real_csv_data(file_path)
            if excel_path and image_path:
                success_count += 1
                excel_files.append(excel_path)
                image_files.append(image_path)
                if data_changes:
                    all_changes.extend(data_changes)
        
        print(f"\n🎉 === 处理完成 ===")
        print(f"📊 成功处理 {success_count}/{len(mt4_files)} 个MT4文件")
        
        if success_count > 0:
            print(f"✅ 所有文件已按照完整流程处理:")
            print(f"   1️⃣  ✅ 真实CSV数据读取完成")
            print(f"   2️⃣  ✅ Excel文件生成并颜色标记完成") 
            print(f"   3️⃣  ✅ A1:M40截图生成完成")
            
            # 发送邮件报告
            if self.enable_email and self.recipients:
                print(f"\n📧 === 发送邮件报告 ===")
                try:
                    success = self.email_sender.send_mt4_report(
                        excel_files=excel_files,
                        image_files=image_files,
                        recipients=self.recipients,
                        subject=f"MT4数据处理报告 - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                        data_changes=all_changes
                    )
                    if success:
                        print(f"   4️⃣  ✅ 邮件报告发送完成")
                    else:
                        print(f"   4️⃣  ❌ 邮件报告发送失败")
                except Exception as e:
                    print(f"   4️⃣  ❌ 邮件发送失败: {str(e)}")
            elif self.enable_email and not self.recipients:
                print(f"\n⚠️  邮件功能已启用但未设置收件人")
            
            print(f"\n🎉 所有任务已完成！")

def main():
    """
    主函数
    """
    import argparse
    
    parser = argparse.ArgumentParser(description='真实MT4数据处理程序')
    parser.add_argument('--path', type=str, help='MT4数据路径')
    parser.add_argument('--string', type=str, default='KVBt_@_D1', help='目标字符串')
    parser.add_argument('--minutes', type=int, default=5, help='时间限制（分钟）')
    parser.add_argument('--email', action='store_true', help='启用邮件发送功能')
    parser.add_argument('--recipients', type=str, nargs='+', help='收件人邮箱地址（多个用空格分隔）')
    
    args = parser.parse_args()
    
    # 创建并运行处理器
    processor = RealMT4DataProcessor(
        mt4_path=args.path,
        target_string=args.string,
        time_limit_minutes=args.minutes,
        enable_email=args.email,
        recipients=args.recipients
    )
    
    processor.run()

if __name__ == "__main__":
    main()