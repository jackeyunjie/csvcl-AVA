#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
调试单元格颜色信息
"""

import os
from openpyxl import load_workbook

def debug_cell_colors():
    """
    调试单元格颜色信息
    """
    print("=== 单元格颜色调试 ===")
    
    # 查找最新的colored.xlsx文件
    xlsx_files = [f for f in os.listdir('.') if f.endswith('_colored.xlsx')]
    if not xlsx_files:
        print("❌ 未找到Excel文件")
        return
    
    # 选择有最多颜色标记的文件
    target_file = None
    for file in xlsx_files:
        if 'range' in file:
            target_file = file
            break
    
    if not target_file:
        target_file = xlsx_files[0]
    
    print(f"📁 检查文件: {target_file}")
    
    try:
        wb = load_workbook(target_file)
        ws = wb.active
        
        print(f"\n🔍 检查E2:G40区域的颜色信息...")
        
        found_colored_cells = 0
        
        # 检查E2:G40区域
        for row in range(2, 41):
            for col in range(5, 8):  # E, F, G列
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value is not None:
                    # 检查填充颜色
                    fill = cell.fill
                    
                    print(f"\n单元格 {chr(64+col)}{row} = {cell_value}")
                    print(f"  Fill类型: {type(fill)}")
                    print(f"  Fill对象: {fill}")
                    
                    if hasattr(fill, 'start_color'):
                        print(f"  start_color: {fill.start_color}")
                        if hasattr(fill.start_color, 'rgb'):
                            print(f"  RGB值: {fill.start_color.rgb}")
                            if fill.start_color.rgb and fill.start_color.rgb != 'FFFFFF':
                                found_colored_cells += 1
                        if hasattr(fill.start_color, 'index'):
                            print(f"  索引值: {fill.start_color.index}")
                    
                    if hasattr(fill, 'fill_type'):
                        print(f"  填充类型: {fill.fill_type}")
                    
                    # 检查字体颜色
                    font = cell.font
                    print(f"  字体颜色: {font.color}")
                    
                    if found_colored_cells >= 5:  # 只显示前5个有颜色的单元格
                        break
            if found_colored_cells >= 5:
                break
        
        print(f"\n✅ 找到 {found_colored_cells} 个有颜色的单元格")
        
    except Exception as e:
        print(f"❌ 错误: {str(e)}")

if __name__ == "__main__":
    debug_cell_colors()