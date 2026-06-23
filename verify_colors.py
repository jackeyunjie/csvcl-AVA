#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证颜色标注是否正确显示
"""

import os
import sys
from PIL import Image, ImageDraw, ImageFont
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

def analyze_excel_colors(excel_file):
    """分析Excel文件中的颜色标注"""
    print(f"分析Excel文件: {excel_file}")
    
    if not os.path.exists(excel_file):
        print(f"❌ 文件不存在: {excel_file}")
        return
    
    # 加载Excel工作簿
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # 定义颜色映射
        color_map = {
            'FF0000': '红色',
            'FFCCCC': '淡红色',
            '00FF00': '绿色',
            'CCFFCC': '淡绿色',
            'FFFF00': '黄色'
        }
        
        # 分析E2:G40区域的颜色
        color_stats = {}
        for row in range(2, 41):
            for col in range(5, 8):  # E=5, F=6, G=7
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value is not None:
                    # 获取单元格填充颜色
                    fill_color = None
                    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                        fill_color = cell.fill.start_color.rgb
                    
                    if fill_color and fill_color != '00000000':  # 不是默认颜色
                        # 去掉前两个字符（如果是00开头）
                        if fill_color.startswith('00') and len(fill_color) > 6:
                            fill_color = fill_color[2:]
                        
                        color_name = color_map.get(fill_color, f"未知({fill_color})")
                        
                        # 更新统计
                        if color_name not in color_stats:
                            color_stats[color_name] = 0
                        color_stats[color_name] += 1
                        
                        col_letter = chr(ord('A') + col - 1)
                        print(f"单元格 {col_letter}{row} 值={cell_value} 颜色={color_name}")
        
        # 打印统计结果
        print("\n颜色统计:")
        for color, count in color_stats.items():
            print(f"  {color}: {count}个单元格")
        
        return color_stats
    
    except Exception as e:
        print(f"❌ 分析Excel文件时出错: {str(e)}")
        return None

def analyze_image_colors(image_file):
    """分析图片中的颜色"""
    print(f"\n分析图片文件: {image_file}")
    
    if not os.path.exists(image_file):
        print(f"❌ 文件不存在: {image_file}")
        return
    
    try:
        # 打开图片
        img = Image.open(image_file)
        
        # 显示图片信息
        print(f"图片尺寸: {img.size}")
        print(f"图片模式: {img.mode}")
        
        # 转换为RGB模式（如果不是）
        if img.mode != 'RGB':
            img = img.convert('RGB')
        
        # 定义颜色范围（RGB值）
        color_ranges = {
            '红色': ((200, 0, 0), (255, 100, 100)),
            '淡红色': ((200, 150, 150), (255, 220, 220)),
            '绿色': ((0, 200, 0), (100, 255, 100)),
            '淡绿色': ((150, 200, 150), (220, 255, 220)),
            '黄色': ((200, 200, 0), (255, 255, 100))
        }
        
        # 分析图片中的颜色
        color_pixels = {color: 0 for color in color_ranges}
        
        # 获取图片数据
        img_data = np.array(img)
        
        # 对每种颜色进行分析
        for color_name, ((r_min, g_min, b_min), (r_max, g_max, b_max)) in color_ranges.items():
            # 创建颜色掩码
            mask = (
                (img_data[:,:,0] >= r_min) & (img_data[:,:,0] <= r_max) &
                (img_data[:,:,1] >= g_min) & (img_data[:,:,1] <= g_max) &
                (img_data[:,:,2] >= b_min) & (img_data[:,:,2] <= b_max)
            )
            
            # 计算像素数量
            color_pixels[color_name] = np.sum(mask)
        
        # 打印统计结果
        print("\n图片颜色统计:")
        for color, count in color_pixels.items():
            print(f"  {color}: 约 {count} 像素")
        
        # 显示图片（可选）
        plt.figure(figsize=(10, 8))
        plt.imshow(img)
        plt.title("图片颜色分析")
        plt.axis('off')
        plt.show()
        
        return color_pixels
    
    except Exception as e:
        print(f"❌ 分析图片时出错: {str(e)}")
        return None

def main():
    """主函数"""
    print("=== 颜色显示验证工具 ===")
    
    # 查找最新的Excel文件和图片
    excel_files = glob.glob("*_colored.xlsx")
    image_files = glob.glob("*_full_table.png")
    
    if not excel_files:
        print("❌ 未找到Excel文件")
        return
    
    if not image_files:
        print("❌ 未找到图片文件")
        return
    
    # 按修改时间排序，获取最新的文件
    latest_excel = max(excel_files, key=os.path.getmtime)
    latest_image = max(image_files, key=os.path.getmtime)
    
    print(f"最新Excel文件: {latest_excel}")
    print(f"最新图片文件: {latest_image}")
    
    # 分析Excel文件中的颜色
    excel_colors = analyze_excel_colors(latest_excel)
    
    # 分析图片中的颜色
    image_colors = analyze_image_colors(latest_image)
    
    # 比较结果
    if excel_colors and image_colors:
        print("\n=== 比较结果 ===")
        print("Excel文件和图片中的颜色应该大致匹配")
        print("如果图片中缺少某些颜色，可能是图片生成过程中出现了问题")

if __name__ == "__main__":
    import glob
    main()